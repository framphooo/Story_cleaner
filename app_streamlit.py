"""
Streamlit web app for spreadsheet normalization.

This app wraps the existing normalization engine (clean_for_snowflake.py)
to provide a user-friendly interface for processing Excel files.
"""

import streamlit as st
import streamlit.components.v1 as components
from pathlib import Path
import json
from datetime import datetime, timedelta
import tempfile
import shutil
import pandas as pd
import time
import zipfile
from io import BytesIO
import hashlib
import os
import atexit

# Import the normalization function
from clean_for_snowflake import normalize_spreadsheet

# ============================================================================
# SECURITY MODULE: Authentication, Rate Limiting, and File Cleanup
# ============================================================================

# Password configuration - CHANGE THIS PASSWORD!
# Priority: Streamlit secrets > Environment variable > Default
# For Streamlit Cloud: Use Secrets (Settings → Secrets)
# For local: Use environment variable: export APP_PASSWORD="your_password"
try:
    # Try Streamlit secrets first (for Streamlit Cloud)
    APP_PASSWORD = st.secrets.get("security", {}).get("password", None)
    if APP_PASSWORD is None:
        # Fall back to environment variable
        APP_PASSWORD = os.getenv("APP_PASSWORD", "ChangeMe123!")
except (AttributeError, FileNotFoundError, KeyError):
    # If secrets not available, use environment variable or default
    APP_PASSWORD = os.getenv("APP_PASSWORD", "ChangeMe123!")

# Rate limiting configuration
MAX_REQUESTS_PER_HOUR = 20  # Maximum processing requests per hour per session
RATE_LIMIT_WINDOW = 3600  # 1 hour in seconds

def hash_password(password: str) -> str:
    """Hash password using SHA256."""
    return hashlib.sha256(password.encode()).hexdigest()

def check_password(password: str, password_hash: str) -> bool:
    """Verify password against hash."""
    return hash_password(password) == password_hash

def check_rate_limit() -> tuple[bool, str]:
    """Check if user has exceeded rate limit. Returns (allowed, message)."""
    if 'rate_limit_requests' not in st.session_state:
        st.session_state.rate_limit_requests = []
    
    now = time.time()
    # Remove requests older than the rate limit window
    st.session_state.rate_limit_requests = [
        req_time for req_time in st.session_state.rate_limit_requests
        if now - req_time < RATE_LIMIT_WINDOW
    ]
    
    # Check if limit exceeded
    if len(st.session_state.rate_limit_requests) >= MAX_REQUESTS_PER_HOUR:
        remaining_time = RATE_LIMIT_WINDOW - (now - st.session_state.rate_limit_requests[0])
        minutes = int(remaining_time / 60)
        return False, f"Rate limit exceeded. Please wait {minutes} minutes before processing more files."
    
    return True, ""

def record_request():
    """Record a processing request for rate limiting."""
    if 'rate_limit_requests' not in st.session_state:
        st.session_state.rate_limit_requests = []
    st.session_state.rate_limit_requests.append(time.time())

def cleanup_session_files():
    """Clean up all temporary files from current session."""
    try:
        # Clean up temp_jobs directories created in this session
        if 'session_job_dirs' in st.session_state:
            for job_dir in st.session_state.session_job_dirs:
                if job_dir.exists():
                    try:
                        shutil.rmtree(job_dir)
                    except Exception as e:
                        st.error(f"Error cleaning up {job_dir}: {e}")
            st.session_state.session_job_dirs = []
        
        # Clean up any temp files in temp_jobs older than 1 hour
        temp_jobs_dir = Path("temp_jobs")
        if temp_jobs_dir.exists():
            now = time.time()
            for item in temp_jobs_dir.iterdir():
                if item.is_dir():
                    try:
                        # Check if directory is older than 1 hour
                        mtime = item.stat().st_mtime
                        if now - mtime > 3600:  # 1 hour
                            shutil.rmtree(item)
                    except Exception:
                        pass  # Ignore errors for old cleanup
    except Exception as e:
        # Don't show errors to user for background cleanup
        pass

def secure_delete_file(file_path: Path):
    """Securely delete a file by overwriting and removing."""
    try:
        if file_path.exists():
            # Overwrite with zeros (simple secure delete)
            if file_path.is_file():
                with open(file_path, "ba+", buffering=0) as f:
                    f.seek(0)
                    f.write(b'\x00' * min(1024, file_path.stat().st_size))
                file_path.unlink()
            elif file_path.is_dir():
                shutil.rmtree(file_path)
    except Exception:
        pass  # Ignore errors during cleanup

def authenticate():
    """Handle user authentication. Returns True if authenticated."""
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
    
    if not st.session_state.authenticated:
        st.markdown("""
        <style>
        /* Load Poppins font */
        @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@400;600&display=swap');
        
        /* Hide everything Streamlit default */
        header[data-testid="stHeader"],
        #MainMenu,
        footer,
        .stDeployButton {
            display: none !important;
        }
        
        /* Full page dark green background */
        .stApp {
            background-color: #0D3A32 !important;
        }
        
        /* Remove all padding and margins */
        .main .block-container {
            padding: 0 !important;
            margin: 0 !important;
            max-width: 100% !important;
        }
        
        /* Prevent body scroll */
        body {
            overflow: hidden !important;
        }
        
        /* Center login content vertically and horizontally - fixed viewport */
        .login-wrapper {
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            display: flex;
            justify-content: center;
            align-items: center;
            padding: 2rem;
        }
        
        .login-simple {
            text-align: center;
            color: #FFFFFF !important;
            width: 100%;
            max-width: 400px;
        }
        
        /* Ensure all text is white and uses Poppins */
        .login-simple,
        .login-simple *,
        .login-simple h3,
        .login-simple p,
        .login-simple div,
        .login-simple h3 * {
            color: #FFFFFF !important;
            font-family: 'Poppins', sans-serif !important;
        }
        
        /* Override any Streamlit default text colors */
        .stMarkdown,
        .stMarkdown h3,
        .stMarkdown p {
            color: #FFFFFF !important;
        }
        
        /* Style placeholder text */
        .stTextInput > div > div > input::placeholder {
            color: #999999 !important;
            font-family: 'Poppins', sans-serif !important;
        }
        
        /* Button text should be dark (on teal background) */
        .stButton > button {
            color: #1B1B1B !important;
        }
        
        /* Style password input */
        .stTextInput > div > div > input {
            background-color: #FFFFFF !important;
            color: #1B1B1B !important;
            border-radius: 12px !important;
            padding: 0.75rem 1rem !important;
            font-family: 'Poppins', sans-serif !important;
            width: 100% !important;
            margin: 0 auto 1.5rem auto !important;
        }
        
        /* Style button */
        .stButton > button {
            background-color: #00FDCF !important;
            color: #1B1B1B !important;
            font-weight: 600 !important;
            border-radius: 12px !important;
            border: none !important;
            padding: 0.75rem 2rem !important;
            font-family: 'Poppins', sans-serif !important;
            width: 100% !important;
            margin: 0 auto !important;
        }
        
        .stButton > button:hover {
            background-color: #00E6BF !important;
        }
        
        /* Error message styling */
        .stAlert {
            max-width: 100%;
            margin: 1rem auto !important;
        }
        </style>
        """, unsafe_allow_html=True)
        
        st.markdown('<div class="login-wrapper"><div class="login-simple">', unsafe_allow_html=True)
        st.markdown("""
        <div style="color: #FFFFFF !important; font-family: 'Poppins', sans-serif !important;">
            <h3 style="color: #FFFFFF !important; font-family: 'Poppins', sans-serif !important; font-weight: 600 !important; 
                       font-size: 1.75rem !important; margin-bottom: 1rem !important; letter-spacing: 0.5px !important;
                       line-height: 1.4 !important;">
                🔒 Enter password to access
            </h3>
            <p style="color: #FFFFFF !important; font-family: 'Poppins', sans-serif !important; font-weight: 400 !important; 
                      font-size: 1rem !important; margin-bottom: 2.5rem !important; opacity: 0.95 !important;
                      line-height: 1.5 !important;">
                Secure access to your spreadsheet normalizer
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        password = st.text_input("", type="password", key="auth_password", label_visibility="collapsed", placeholder="Password")
        
        if st.button("Access", type="primary", use_container_width=True):
            stored_hash = hash_password(APP_PASSWORD)
            if check_password(password, stored_hash):
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Incorrect password")
        
        st.markdown('</div></div>', unsafe_allow_html=True)
        st.stop()
    
    return True

# Page configuration
st.set_page_config(
    page_title="Spreadsheet Normalizer",
    page_icon="📊",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# ============================================================================
# AUTHENTICATION - Must be first thing after page config
# ============================================================================
authenticate()

# Initialize session state for file tracking
if 'session_job_dirs' not in st.session_state:
    st.session_state.session_job_dirs = []

# Register cleanup function to run on session end
if 'cleanup_registered' not in st.session_state:
    atexit.register(cleanup_session_files)
    st.session_state.cleanup_registered = True


# Custom expander using HTML details/summary - completely bypasses Streamlit's broken expander
def custom_expander(label, expanded=False, content_html=None):
    """Custom expander using native HTML details/summary - no Material Icons, arrow emoji on right.
    
    Args:
        label: The label for the expander (will get emoji prefix based on label text)
        expanded: Whether the expander starts expanded
        content_html: If provided, renders this HTML as content. Otherwise returns context manager.
    """
    import uuid
    from html import escape
    
    # Add emoji based on label
    emoji_prefix = ""
    if "Info" in label or "info" in label.lower():
        emoji_prefix = "ℹ️ "
    elif "Warning" in label or "warning" in label.lower():
        emoji_prefix = "⚠️ "
    elif "Error" in label or "error" in label.lower():
        emoji_prefix = "❌ "
    
    expander_id = f"custom_exp_{uuid.uuid4().hex[:8]}"
    arrow_emoji = '⬆️' if expanded else '⬇️'
    
    # Inject CSS for custom expanders (only once)
    if 'custom_expander_css_injected' not in st.session_state:
        st.markdown("""
        <style>
        .custom-expander-details {
            margin: 1rem 0;
            border: 1px solid #EEEEEE;
            border-radius: 12px;
            background-color: #FFFFFF;
            overflow: hidden;
        }
        .custom-expander-summary {
            padding: 0.75rem 1rem;
            cursor: pointer;
            font-weight: 600;
            font-family: 'Poppins', sans-serif;
            color: #1B1B1B;
            background-color: #FFFFFF;
            list-style: none;
            display: flex;
            justify-content: space-between;
            align-items: center;
            transition: background-color 0.2s;
        }
        .custom-expander-summary:hover {
            background-color: #F8F8F8;
        }
        .custom-expander-summary::-webkit-details-marker {
            display: none;
        }
        .custom-expander-summary::marker {
            display: none;
        }
        .custom-expander-arrow {
            font-size: 1rem;
            margin-left: 1rem;
            transition: transform 0.2s;
        }
        .custom-expander-details[open] .custom-expander-arrow {
            transform: rotate(180deg);
        }
        .custom-expander-content {
            padding: 1rem;
            border-top: 1px solid #EEEEEE;
            background-color: #FFFFFF;
            font-family: 'Poppins', sans-serif;
        }
        </style>
        """, unsafe_allow_html=True)
        st.session_state.custom_expander_css_injected = True
    
    # If content_html provided, render directly
    if content_html is not None:
        st.markdown(f"""
        <details {'open' if expanded else ''} class="custom-expander-details" id="{expander_id}">
            <summary class="custom-expander-summary">
                <span>{emoji_prefix}{escape(label)}</span>
                <span class="custom-expander-arrow">{arrow_emoji}</span>
            </summary>
            <div class="custom-expander-content">
                {content_html}
            </div>
        </details>
        """, unsafe_allow_html=True)
        return None
    
    # Return context manager for content
    class CustomExpander:
        def __init__(self, label, expanded, expander_id, arrow, emoji):
            self.label = label
            self.expanded = expanded
            self.expander_id = expander_id
            self.arrow = arrow
            self.emoji = emoji
            self.content_parts = []
            self._inside_context = False
        
        def __enter__(self):
            self._inside_context = True
            # Store where we'll insert content
            self.content_placeholder = st.empty()
            # Render the opening tag
            st.markdown(f"""
            <details {'open' if self.expanded else ''} class="custom-expander-details" id="{self.expander_id}">
                <summary class="custom-expander-summary">
                    <span>{self.emoji}{self.label}</span>
                    <span class="custom-expander-arrow">{self.arrow}</span>
                </summary>
                <div class="custom-expander-content">
            """, unsafe_allow_html=True)
            return self
        
        def markdown(self, text):
            """Add markdown content to the expander."""
            if self._inside_context:
                st.markdown(text)
        
        def __exit__(self, *args):
            st.markdown("</div></details>", unsafe_allow_html=True)
            self._inside_context = False
    
    return CustomExpander(label, expanded, expander_id, arrow_emoji, emoji_prefix)

# CSS injection for Poppins font and brand styling
# Version: 2024-12-14-v2 - Fixed expander arrows
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@400;600&display=swap');
    
    * {
        font-family: 'Poppins', sans-serif !important;
    }
    
    h1, h2, h3, h4, h5, h6 {
        font-weight: 600 !important;
        color: #1B1B1B;
    }
    
    /* Reduce spacing after title */
    h1 {
        margin-bottom: 0.5rem !important;
    }
    
    /* Reduce spacing around horizontal rules */
    hr {
        margin-top: 0.5rem !important;
        margin-bottom: 0.5rem !important;
    }
    
    /* COMPLETE REMOVAL: Hide and disable all anchor links that Streamlit auto-generates for headers */
    h1 a, h2 a, h3 a, h4 a, h5 a, h6 a {
        display: none !important;
        visibility: hidden !important;
        width: 0 !important;
        height: 0 !important;
        opacity: 0 !important;
        pointer-events: none !important;
        position: absolute !important;
        left: -9999px !important;
    }
    
    /* Also hide the anchor icon/link that Streamlit adds to headers */
    .stMarkdown h1 a,
    .stMarkdown h2 a,
    .stMarkdown h3 a,
    .stMarkdown h4 a,
    .stMarkdown h5 a,
    .stMarkdown h6 a {
        display: none !important;
        visibility: hidden !important;
        pointer-events: none !important;
    }
    
    /* Hide anchor links in info cards */
    .info-card h3 a {
        display: none !important;
        visibility: hidden !important;
        pointer-events: none !important;
    }
    
    /* Prevent headers from being clickable or creating anchor links */
    h1, h2, h3, h4, h5, h6 {
        cursor: default !important;
    }
    
    /* Remove any anchor IDs that Streamlit might add to headers */
    h1[id], h2[id], h3[id], h4[id], h5[id], h6[id] {
        scroll-margin-top: 0 !important;
    }
    
    :root {
        --bg-color: #F3F1EC;
        --text-color: #1B1B1B;
        --card-bg: #FFFFFF;
    }
    
    body {
        background-color: var(--bg-color);
    }
    
    
    .stButton>button {
        background-color: #00FDCF;
        color: #1B1B1B;
        font-weight: 600;
        border-radius: 12px;
        border: none;
        padding: 0.5rem 2rem;
        transition: all 0.3s;
    }
    
    .stButton>button:hover:not(:disabled) {
        background-color: #00E6BF;
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0, 253, 207, 0.3);
    }
    
    .stButton>button:disabled {
        background-color: #EEEEEE !important;
        color: #999999 !important;
        cursor: not-allowed !important;
        opacity: 0.6;
    }
    
    .stFileUploader>div>div>div>div {
        border-radius: 12px;
        border: 2px dashed #00FDCF;
        background-color: #FFFFFF;
    }
    
    .main .block-container {
        max-width: 900px;
        padding-top: 2rem;
    }
    
    .info-box {
        background-color: #D6F5F0;
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0;
        border-left: 4px solid #00FDCF;
    }
    
    .status-success {
        background-color: transparent;
        padding: 0;
        margin: 1rem 0;
    }
    
    .status-warning {
        background-color: transparent;
        padding: 0;
        margin: 1rem 0;
    }
    
    .status-error {
        background-color: transparent;
        padding: 0;
        margin: 1rem 0;
    }
    
    .processing-overlay {
        position: fixed;
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        background-color: rgba(243, 241, 236, 0.95);
        display: flex;
        justify-content: center;
        align-items: center;
        z-index: 9999;
        flex-direction: column;
        pointer-events: none;
    }
    
    .processing-overlay.showing {
        pointer-events: auto;
    }
    
    .processing-card {
        background-color: #FFFFFF;
        padding: 3rem;
        border-radius: 14px;
        box-shadow: 0 8px 24px rgba(0, 0, 0, 0.1);
        text-align: center;
        max-width: 400px;
    }
    
    .progress-ring-container {
        width: 120px;
        height: 120px;
        margin: 0 auto 1.5rem;
        position: relative;
    }
    
    .progress-ring {
        width: 120px;
        height: 120px;
        transform: rotate(-90deg);
    }
    
    .progress-ring-circle {
        fill: none;
        stroke-width: 8;
        stroke-linecap: round;
    }
    
    .progress-ring-bg {
        stroke: #EEEEEE;
    }
    
    .progress-ring-progress {
        stroke: #00FDCF;
        transition: stroke-dashoffset 0.3s;
    }
    
    .check-icon {
        position: absolute;
        top: 50%;
        left: 50%;
        transform: translate(-50%, -50%);
        font-size: 48px;
        color: #FFFFFF;
        font-weight: bold;
        opacity: 0;
        transition: opacity 0.3s;
    }
    
    .check-icon.visible {
        opacity: 1;
    }
    
    .processing-text {
        font-size: 1.2rem;
        font-weight: 600;
        color: #1B1B1B;
        margin-bottom: 0.5rem;
    }
    
    .done-text {
        font-size: 1rem;
        color: #0D3A32;
        margin-top: 1rem;
    }
    
    .summary-card {
        background-color: #FFFFFF;
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0;
        border-left: 4px solid #00FDCF;
    }
    
    .summary-bullet {
        margin: 0.5rem 0;
        color: #1B1B1B;
    }
    
    .header-link {
        color: #00FDCF;
        text-decoration: none;
        font-weight: 600;
        cursor: pointer;
        margin-left: 1rem;
    }
    
    .header-link:hover {
        text-decoration: underline;
    }
    
    .run-history-item {
        background-color: #FFFFFF;
        padding: 1rem;
        border-radius: 12px;
        margin: 0.5rem 0;
        border-left: 4px solid #EEEEEE;
    }
    
    .run-history-item.success {
        border-left-color: #00FDCF;
    }
    
    .run-history-item.warning {
        border-left-color: #FFA500;
    }
    
    .run-history-item.error {
        border-left-color: #FF0000;
    }
    
    .feedback-modal {
        display: none;
        position: fixed;
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        background-color: rgba(0, 0, 0, 0.5);
        z-index: 10000;
        justify-content: center;
        align-items: center;
    }
    
    .feedback-modal.active {
        display: flex;
    }
    
    .feedback-modal-content {
        background-color: #FFFFFF;
        padding: 2rem;
        border-radius: 14px;
        max-width: 500px;
        width: 90%;
        box-shadow: 0 8px 24px rgba(0, 0, 0, 0.2);
        position: relative;
    }
    
    .feedback-modal-close {
        position: absolute;
        top: 1rem;
        right: 1rem;
        background: none;
        border: none;
        font-size: 1.5rem;
        cursor: pointer;
        color: #1B1B1B;
        font-weight: bold;
        line-height: 1;
    }
    
    .feedback-modal-close:hover {
        color: #00FDCF;
    }
    
    .feedback-modal-title {
        font-size: 1.5rem;
        font-weight: 600;
        color: #1B1B1B;
        margin-bottom: 1rem;
        margin-top: 0;
    }
    
    .feedback-small-button {
        background-color: #00FDCF;
        color: #1B1B1B;
        font-weight: 600;
        border-radius: 8px;
        border: none;
        padding: 0.4rem 1rem;
        font-size: 0.9rem;
        cursor: pointer;
        transition: all 0.3s;
    }
    
    .feedback-small-button:hover {
        background-color: #00E6BF;
        transform: translateY(-1px);
    }
    
    .progress-ring-progress-conic {
        background: conic-gradient(#00FDCF 0% var(--progress), #EEEEEE var(--progress) 100%);
        border-radius: 50%;
        width: 120px;
        height: 120px;
        display: flex;
        align-items: center;
        justify-content: center;
        margin: 0 auto 1.5rem;
        transition: background 0.3s ease;
        position: relative;
    }
    
    .progress-ring-progress-conic.completed {
        background: #00FDCF !important;
        background-color: #00FDCF !important;
    }
    
    .progress-ring-inner {
        background-color: #FFFFFF;
        border-radius: 50%;
        width: 100px;
        height: 100px;
        display: flex;
        align-items: center;
        justify-content: center;
        position: relative;
        z-index: 1;
    }
    
    .progress-ring-inner .check-icon {
        font-size: 48px;
        color: #FFFFFF !important;
        font-weight: bold;
        opacity: 0;
        transition: opacity 0.3s ease;
        position: relative;
        z-index: 2;
        text-shadow: 0 0 2px rgba(0, 0, 0, 0.2);
    }
    
    .progress-ring-inner .check-icon.visible {
        opacity: 1 !important;
        display: flex !important;
    }
    
    /* Expander header styling - fix for Material Icons overlap issue */
    .streamlit-expanderHeader {
        font-weight: 600;
        position: relative;
        padding-right: 2rem !important;
        cursor: pointer;
        overflow: hidden !important;
        font-family: 'Poppins', sans-serif !important;
    }
    
    /* Ensure Poppins font is used everywhere in expander headers */
    .streamlit-expanderHeader,
    .streamlit-expanderHeader label,
    .streamlit-expanderHeader span {
        font-family: 'Poppins', sans-serif !important;
    }
    
    /* COMPLETE HIDE: All Material Icons elements - using multiple strategies */
    .streamlit-expanderHeader i,
    .streamlit-expanderHeader i[class*="material"],
    .streamlit-expanderHeader [class*="material-icons"],
    .streamlit-expanderHeader [class*="MaterialIcons"],
    .streamlit-expanderHeader [class*="material"],
    .streamlit-expanderHeader svg,
    .streamlit-expanderHeader path,
    .streamlit-expanderHeader g {
        display: none !important;
        visibility: hidden !important;
        width: 0 !important;
        height: 0 !important;
        opacity: 0 !important;
        font-size: 0 !important;
        line-height: 0 !important;
        margin: 0 !important;
        padding: 0 !important;
        position: absolute !important;
        left: -9999px !important;
        overflow: hidden !important;
        text-indent: -9999px !important;
        clip-path: inset(100%) !important;
        clip: rect(0, 0, 0, 0) !important;
    }
    
    /* Custom arrow on the RIGHT side using CSS ::after pseudo-element */
    .streamlit-expanderHeader::after {
        content: '▼' !important;
        position: absolute !important;
        right: 0.75rem !important;
        top: 50% !important;
        transform: translateY(-50%) !important;
        font-style: normal !important;
        font-size: 0.85rem !important;
        font-weight: normal !important;
        pointer-events: none !important;
        font-family: 'Poppins', sans-serif !important;
        color: #1B1B1B !important;
        z-index: 1000 !important;
        display: inline-block !important;
        visibility: visible !important;
        opacity: 1 !important;
        line-height: 1 !important;
        width: auto !important;
        height: auto !important;
        margin: 0 !important;
        padding: 0 !important;
        transition: transform 0.2s ease !important;
    }
    
    /* Arrow points UP when expanded */
    .streamlit-expanderHeader[aria-expanded="true"]::after {
        content: '▲' !important;
    }
    
    
    /* JavaScript-created arrow span (backup solution) */
    .streamlit-expanderHeader .custom-expander-arrow {
        position: absolute !important;
        right: 0.75rem !important;
        top: 50% !important;
        transform: translateY(-50%) !important;
        font-size: 0.85rem !important;
        font-weight: normal !important;
        pointer-events: none !important;
        font-family: 'Poppins', sans-serif !important;
        color: #1B1B1B !important;
        z-index: 1000 !important;
        display: inline-block !important;
        visibility: visible !important;
        opacity: 1 !important;
        line-height: 1 !important;
    }
    
    
    /* Status badge styles */
    .status-badge {
        display: inline-block;
        padding: 0.5rem 1rem;
        border-radius: 12px;
        font-size: 0.9rem;
        font-weight: 400;
        text-align: left;
        line-height: 1.5;
        width: 100%;
        box-sizing: border-box;
    }
    
    .status-badge strong {
        font-weight: 600;
        margin-right: 0.5rem;
    }
    
    .status-badge.completed {
        background-color: #D6F5F0;
        color: #0D3A32;
        border: 1px solid #00FDCF;
    }
    
    .status-badge.warnings {
        background-color: #FFF4E6;
        color: #8B6914;
        border: 1px solid #FFA500;
    }
    
    .status-badge.error {
        background-color: #FFE6E6;
        color: #8B0000;
        border: 1px solid #FF0000;
    }
    
    /* Info card styles */
    .info-card {
        background-color: #FFFFFF;
        padding: 1.5rem;
        border-radius: 12px;
        margin: 1rem 0;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.05);
        border-left: 4px solid #00FDCF;
        transition: all 0.3s ease;
        animation: fadeInUp 0.5s ease-out;
    }
    
    .info-card:hover {
        box-shadow: 0 4px 12px rgba(0, 253, 207, 0.15);
        transform: translateY(-2px);
    }
    
    .info-card h3 {
        margin-top: 0;
        color: #1B1B1B;
        font-weight: 600;
    }
    
    @keyframes fadeInUp {
        from {
            opacity: 0;
            transform: translateY(10px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    /* Animated Data Particles - Side decorations */
    .data-particles-container {
        position: fixed;
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        pointer-events: none;
        z-index: 1;
        overflow: hidden;
    }
    
    .data-particle {
        position: absolute;
        width: 8px;
        height: 8px;
        background-color: #00FDCF;
        border-radius: 50%;
        opacity: 0;
        pointer-events: none;
        box-shadow: 0 0 8px rgba(0, 253, 207, 0.8);
    }
    
    .data-particle.left {
        left: 5%;
        animation: particleFloatLeft 8s infinite ease-in-out;
    }
    
    .data-particle.right {
        right: 5%;
        animation: particleFloatRight 8s infinite ease-in-out;
    }
    
    @keyframes particleFloatLeft {
        0% {
            opacity: 0;
            transform: translateY(100vh) translateX(-20px) scale(0.5);
        }
        10% {
            opacity: 0.7;
        }
        50% {
            opacity: 0.9;
            transform: translateY(50vh) translateX(0px) scale(1);
        }
        90% {
            opacity: 0.6;
        }
        100% {
            opacity: 0;
            transform: translateY(-10vh) translateX(20px) scale(0.5);
        }
    }
    
    @keyframes particleFloatRight {
        0% {
            opacity: 0;
            transform: translateY(100vh) translateX(20px) scale(0.5);
        }
        10% {
            opacity: 0.7;
        }
        50% {
            opacity: 0.9;
            transform: translateY(50vh) translateX(0px) scale(1);
        }
        90% {
            opacity: 0.6;
        }
        100% {
            opacity: 0;
            transform: translateY(-10vh) translateX(-20px) scale(0.5);
        }
    }
    
    /* Ensure main content is above particles */
    .main .block-container {
        position: relative;
        z-index: 2;
    }
    
    /* Hide particles on small screens for better mobile experience */
    @media (max-width: 768px) {
        .data-particles-container {
            display: none;
        }
    }
    
</style>
<script>
// AGGRESSIVE FIX: Remove Material Icons text and ensure custom arrows work
// This runs immediately and on all DOM updates
(function() {
    function fixExpanderHeaders() {
        const expanderHeaders = document.querySelectorAll('.streamlit-expanderHeader');
        expanderHeaders.forEach(header => {
            // Step 1: Remove ALL Material Icons elements
            const allChildren = Array.from(header.children);
            allChildren.forEach(el => {
                const tagName = el.tagName.toLowerCase();
                const className = (el.className || '').toString().toLowerCase();
                const textContent = (el.textContent || '').trim().toLowerCase();
                
                // Remove if it's an <i> tag, has material-icons class, or contains icon text
                if (tagName === 'i' || 
                    className.includes('material') ||
                    textContent.includes('keyboard_arrow') ||
                    textContent.includes('expand_more') ||
                    textContent.includes('expand_less') ||
                    textContent.includes('chevron_')) {
                    el.remove();
                }
            });
            
            // Step 2: Remove SVG elements
            header.querySelectorAll('svg').forEach(el => el.remove());
            
            // Step 3: Remove text nodes containing Material Icons names
            const walker = document.createTreeWalker(
                header,
                NodeFilter.SHOW_TEXT,
                null
            );
            
            const textNodesToRemove = [];
            let textNode;
            while (textNode = walker.nextNode()) {
                const text = (textNode.textContent || '').trim();
                if (text === 'keyboard_arrow_right' || 
                    text === 'keyboard_arrow_down' ||
                    text === 'keyboard_arrow_up' ||
                    text === 'expand_more' ||
                    text === 'expand_less' ||
                    (text.includes('keyboard_arrow') && text.length < 25) ||
                    (text.includes('expand_') && text.length < 15) ||
                    (text.includes('chevron_') && text.length < 20)) {
                    textNodesToRemove.push(textNode);
                }
            }
            
            textNodesToRemove.forEach(node => {
                const parent = node.parentElement;
                if (parent && parent !== header) {
                    const parentText = (parent.textContent || '').trim();
                    if (parentText === node.textContent.trim()) {
                        parent.remove();
                    } else {
                        node.remove();
                    }
                } else {
                    node.remove();
                }
            });
            
            // Step 4: Ensure custom arrow span exists and is positioned correctly
            let arrowSpan = header.querySelector('.custom-expander-arrow');
            if (!arrowSpan) {
                arrowSpan = document.createElement('span');
                arrowSpan.className = 'custom-expander-arrow';
                header.style.position = 'relative';
                header.appendChild(arrowSpan);
            }
            
            // Force arrow styling
            arrowSpan.style.cssText = 'position: absolute !important; right: 0.75rem !important; top: 50% !important; transform: translateY(-50%) !important; font-size: 0.85rem !important; font-weight: normal !important; pointer-events: none !important; font-family: "Poppins", sans-serif !important; color: #1B1B1B !important; z-index: 1000 !important; display: inline-block !important; visibility: visible !important; opacity: 1 !important; line-height: 1 !important;';
            
            // Update arrow based on expanded state
            const isExpanded = header.getAttribute('aria-expanded') === 'true';
            arrowSpan.textContent = isExpanded ? '▲' : '▼';
        });
    }
    
    // Run immediately
    fixExpanderHeaders();
    
    // Run on DOM ready
    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', fixExpanderHeaders);
    }
    
    // Run with delays to catch dynamically added expanders
    [10, 50, 100, 200, 500, 1000].forEach(delay => {
        setTimeout(fixExpanderHeaders, delay);
    });
    
    // Watch for new expanders and state changes
    const observer = new MutationObserver(() => {
        fixExpanderHeaders();
    });
    
    observer.observe(document.body, {
        childList: true,
        subtree: true,
        attributes: true,
        attributeFilter: ['aria-expanded', 'class']
    });
    
    // Update arrows on click
    document.addEventListener('click', (e) => {
        if (e.target.closest('.streamlit-expanderHeader')) {
            setTimeout(fixExpanderHeaders, 5);
        }
    }, true);
    
    // Also run on every frame for a short period (aggressive but ensures it works)
    let frameCount = 0;
    const maxFrames = 60; // Run for ~1 second at 60fps
    function frameCheck() {
        if (frameCount < maxFrames) {
            fixExpanderHeaders();
            frameCount++;
            requestAnimationFrame(frameCheck);
        }
    }
    requestAnimationFrame(frameCheck);
})();

// PREVENT URL HASH CHANGES - Keep everything on single URL
// This ensures all sections stay on the same URL without creating separate routes
(function() {
    // Store the original URL without hash
    let baseUrl = window.location.href.split('#')[0];
    
    // Function to remove hash from URL if it appears
    function removeHash() {
        if (window.location.hash) {
            // Replace URL without hash, but don't trigger page reload
            window.history.replaceState(null, null, baseUrl);
        }
    }
    
    // Remove hash immediately if present
    removeHash();
    
    // Update baseUrl if the URL changes (but without hash)
    function updateBaseUrl() {
        const currentUrl = window.location.href.split('#')[0];
        if (currentUrl !== baseUrl) {
            baseUrl = currentUrl;
        }
    }
    
    // Listen for hash changes and remove them immediately
    window.addEventListener('hashchange', function(e) {
        e.preventDefault();
        e.stopPropagation();
        removeHash();
        return false;
    }, true); // Use capture phase to catch early
    
    // Override pushState to prevent hash changes
    const originalPushState = history.pushState;
    history.pushState = function() {
        const args = Array.prototype.slice.call(arguments);
        // If trying to push a hash, remove it
        if (args[2] && typeof args[2] === 'string' && args[2].includes('#')) {
            args[2] = args[2].split('#')[0];
        }
        updateBaseUrl();
        return originalPushState.apply(history, args);
    };
    
    // Override replaceState to prevent hash changes
    const originalReplaceState = history.replaceState;
    history.replaceState = function() {
        const args = Array.prototype.slice.call(arguments);
        // If trying to replace with a hash, remove it
        if (args[2] && typeof args[2] === 'string' && args[2].includes('#')) {
            args[2] = args[2].split('#')[0];
        }
        updateBaseUrl();
        return originalReplaceState.apply(history, args);
    };
    
    // Remove all click handlers from headers that might create anchor links
    function disableHeaderAnchors() {
        const headers = document.querySelectorAll('h1, h2, h3, h4, h5, h6');
        headers.forEach(header => {
            // Remove any anchor links inside headers
            const anchors = header.querySelectorAll('a');
            anchors.forEach(anchor => {
                // Remove href to prevent navigation
                if (anchor.hasAttribute('href')) {
                    anchor.removeAttribute('href');
                }
                // Make anchor non-clickable
                anchor.style.pointerEvents = 'none';
                anchor.style.cursor = 'default';
                // Prevent any click events
                anchor.onclick = function(e) {
                    e.preventDefault();
                    e.stopPropagation();
                    return false;
                };
            });
            
            // Remove any click handlers from headers themselves
            header.onclick = function(e) {
                e.preventDefault();
                e.stopPropagation();
                return false;
            };
            
            // Remove any ID attributes that might be used for anchors
            if (header.id && header.id.trim()) {
                header.removeAttribute('id');
            }
            
            // Remove data attributes that might be used for navigation
            header.removeAttribute('data-testid');
        });
    }
    
    // Run immediately and on DOM changes
    disableHeaderAnchors();
    
    // Watch for new headers being added and remove their anchor capabilities
    const headerObserver = new MutationObserver(function(mutations) {
        let shouldRemoveHash = false;
        mutations.forEach(function(mutation) {
            if (mutation.type === 'attributes') {
                // If an ID or href was added, remove it
                if (mutation.attributeName === 'id' || mutation.attributeName === 'href') {
                    shouldRemoveHash = true;
                }
            }
            if (mutation.type === 'childList' && mutation.addedNodes.length > 0) {
                // New nodes added, check for headers
                mutation.addedNodes.forEach(function(node) {
                    if (node.nodeType === 1) { // Element node
                        if (node.tagName && ['H1', 'H2', 'H3', 'H4', 'H5', 'H6'].includes(node.tagName)) {
                            shouldRemoveHash = true;
                        }
                        // Check children too
                        if (node.querySelectorAll) {
                            const childHeaders = node.querySelectorAll('h1, h2, h3, h4, h5, h6');
                            if (childHeaders.length > 0) {
                                shouldRemoveHash = true;
                            }
                        }
                    }
                });
            }
        });
        
        if (shouldRemoveHash) {
            removeHash();
            disableHeaderAnchors();
        }
    });
    
    headerObserver.observe(document.body, {
        childList: true,
        subtree: true,
        attributes: true,
        attributeFilter: ['id', 'href', 'data-testid']
    });
    
    // Periodically check and remove hash (aggressive approach)
    setInterval(function() {
        removeHash();
        disableHeaderAnchors();
    }, 200);
    
    // Remove hash on any navigation events
    window.addEventListener('popstate', function(e) {
        removeHash();
    });
    
    // Prevent any anchor clicks from changing URL
    document.addEventListener('click', function(e) {
        const target = e.target;
        // Check if clicking on an anchor that would add a hash
        if (target.tagName === 'A' && target.getAttribute('href') && target.getAttribute('href').startsWith('#')) {
            e.preventDefault();
            e.stopPropagation();
            removeHash();
            return false;
        }
        // Check if clicking on a header with an anchor child
        const header = target.closest('h1, h2, h3, h4, h5, h6');
        if (header) {
            const anchor = header.querySelector('a[href^="#"]');
            if (anchor) {
                e.preventDefault();
                e.stopPropagation();
                removeHash();
                return false;
            }
        }
    }, true); // Use capture phase
})();

// Animated Data Particles - Direct injection into main document
(function() {
    function createParticleSystem() {
        // Remove existing container if it exists
        const existing = document.getElementById('data-particles-container');
        if (existing) {
            existing.remove();
        }
        
        // Create container
        const container = document.createElement('div');
        container.id = 'data-particles-container';
        container.className = 'data-particles-container';
        document.body.appendChild(container);
        
        // Number of particles
        const particleCount = 15;
        
        // Create particles with staggered delays
        for (let i = 0; i < particleCount; i++) {
            // Left side particles
            const leftParticle = document.createElement('div');
            leftParticle.className = 'data-particle left';
            const leftDelay = i * 0.6;
            const leftDuration = 7 + Math.random() * 5;
            leftParticle.style.animationDelay = leftDelay + 's';
            leftParticle.style.animationDuration = leftDuration + 's';
            const leftOpacity = 0.7 + Math.random() * 0.3;
            leftParticle.style.opacity = leftOpacity;
            leftParticle.style.backgroundColor = 'rgba(0, 253, 207, ' + leftOpacity + ')';
            leftParticle.style.left = (3 + Math.random() * 4) + '%';
            container.appendChild(leftParticle);
            
            // Right side particles
            const rightParticle = document.createElement('div');
            rightParticle.className = 'data-particle right';
            const rightDelay = i * 0.6 + 0.3;
            const rightDuration = 7 + Math.random() * 5;
            rightParticle.style.animationDelay = rightDelay + 's';
            rightParticle.style.animationDuration = rightDuration + 's';
            const rightOpacity = 0.7 + Math.random() * 0.3;
            rightParticle.style.opacity = rightOpacity;
            rightParticle.style.backgroundColor = 'rgba(0, 253, 207, ' + rightOpacity + ')';
            rightParticle.style.right = (3 + Math.random() * 4) + '%';
            container.appendChild(rightParticle);
        }
    }
    
    // Initialize immediately
    if (document.body) {
        createParticleSystem();
    } else {
        document.addEventListener('DOMContentLoaded', createParticleSystem);
    }
    
    // Recreate on delays to catch Streamlit reruns
    setTimeout(createParticleSystem, 100);
    setTimeout(createParticleSystem, 500);
    setTimeout(createParticleSystem, 1000);
    setTimeout(createParticleSystem, 2000);
    
    // Watch for changes
    const observer = new MutationObserver(function() {
        const container = document.getElementById('data-particles-container');
        if (!container && document.body) {
            setTimeout(createParticleSystem, 100);
        }
    });
    
    if (document.body) {
        observer.observe(document.body, {
            childList: true,
            subtree: false
        });
    }
})();
</script>
""", unsafe_allow_html=True)


# Initialize session state
if 'results' not in st.session_state:
    st.session_state.results = None
if 'job_id' not in st.session_state:
    st.session_state.job_id = None
if 'processing' not in st.session_state:
    st.session_state.processing = False
if 'run_history' not in st.session_state:
    st.session_state.run_history = []
if 'current_file_name' not in st.session_state:
    st.session_state.current_file_name = None
if 'current_output_format' not in st.session_state:
    st.session_state.current_output_format = None
if 'feedback_modal_open' not in st.session_state:
    st.session_state.feedback_modal_open = False
if 'batch_results' not in st.session_state:
    st.session_state.batch_results = None
if 'processing_complete' not in st.session_state:
    st.session_state.processing_complete = False
if 'last_processed_files' not in st.session_state:
    st.session_state.last_processed_files = None

# Helper function to create Run Report Excel
def create_run_report_excel(results, job_id, file_name, output_format, job_dir):
    """Create a business-friendly Excel report with readable sheets."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    report_path = job_dir / f"run_report_{job_id}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # SUMMARY sheet
    ws_summary = wb.create_sheet("SUMMARY")
    ws_summary.append(["Run Report Summary"])
    ws_summary.append([])
    ws_summary.append(["Run ID", job_id])
    ws_summary.append(["Timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append(["File Name", file_name])
    ws_summary.append(["Output Format", output_format])
    ws_summary.append(["Overall Status", results['status'].upper()])
    ws_summary.append([])
    
    # Style header
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A1'].fill = PatternFill(start_color="00FDCF", end_color="00FDCF", fill_type="solid")
    
    # TABLE_OVERVIEW sheet
    ws_overview = wb.create_sheet("TABLE_OVERVIEW")
    if not results['meta_df'].empty:
        # Headers
        headers = ['Table name', 'Rows', 'Columns', 'Quality flag']
        ws_overview.append(headers)
        
        # Style header row
        for cell in ws_overview[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D6F5F0", end_color="D6F5F0", fill_type="solid")
        
        # Data rows
        for _, row in results['meta_df'].iterrows():
            ws_overview.append([
                row['Tab name'],
                int(row['Rows after clean']),
                int(row['Columns after clean']),
                row['Data quality flag']
            ])
    
    # Separate warnings, errors, and info
    warnings_list = []
    errors_list = []
    info_list = []
    
    # Process warnings and classify
    for warning in results['warnings']:
        warning_str = str(warning)
        if 'split into' in warning_str.lower() or 'split sheet' in warning_str.lower():
            info_list.append({'table': 'N/A', 'category': 'INFO', 'message': warning_str})
        else:
            warnings_list.append({'table': 'N/A', 'category': 'WARNING', 'message': warning_str})
    
    # Process errors
    for error in results['errors']:
        errors_list.append({'table': 'N/A', 'category': 'ERROR', 'message': str(error)})
    
    # Add table-specific warnings and errors from meta_df
    if not results['meta_df'].empty:
        for _, row in results['meta_df'].iterrows():
            table_name = row['Tab name']
            
            # Parse warnings
            if pd.notna(row.get('Warnings', '')) and str(row['Warnings']).strip():
                warnings = str(row['Warnings']).split('; ')
                for w in warnings:
                    if w.strip():
                        if 'split into' in w.lower() or 'split sheet' in w.lower():
                            info_list.append({'table': table_name, 'category': 'INFO', 'message': w})
                        else:
                            warnings_list.append({'table': table_name, 'category': 'WARNING', 'message': w})
            
            # Parse errors
            if pd.notna(row.get('Errors', '')) and str(row['Errors']).strip():
                errors = str(row['Errors']).split('; ')
                for e in errors:
                    if e.strip():
                        errors_list.append({'table': table_name, 'category': 'ERROR', 'message': e})
    
    # WARNINGS sheet (only create if there are warnings)
    if warnings_list:
        ws_warnings = wb.create_sheet("WARNINGS")
        ws_warnings.append(['Table name', 'Category', 'Message'])
        for cell in ws_warnings[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        
        for item in warnings_list:
            ws_warnings.append([item['table'], item['category'], item['message']])
    
    # ERRORS sheet (only create if there are errors)
    if errors_list:
        ws_errors = wb.create_sheet("ERRORS")
        ws_errors.append(['Table name', 'Message'])
        for cell in ws_errors[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        for item in errors_list:
            ws_errors.append([item['table'], item['message']])
    
    # INFO sheet (only create if there are info messages)
    if info_list:
        ws_info = wb.create_sheet("INFO")
        ws_info.append(['Table name', 'Category', 'Message'])
        for cell in ws_info[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D6F5F0", end_color="D6F5F0", fill_type="solid")
        
        for item in info_list:
            ws_info.append([item['table'], item['category'], item['message']])
    
    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(report_path)
    return report_path

# Helper function to generate run summary
def generate_run_summary(results, file_name, output_format):
    """Generate a compact summary text for the run."""
    summary_parts = []
    
    # Status
    if results['status'] == 'success':
        summary_parts.append("Normalization output: Completed successfully")
    elif results['status'] == 'partial':
        summary_parts.append("Normalization output: Completed with warnings")
    else:
        summary_parts.append("Normalization output: Failed")
    
    # Tables processed
    if not results['meta_df'].empty:
        table_count = len(results['meta_df'])
        summary_parts.append(f"Tables processed: {table_count}")
        
        # New tables created from splits
        split_count = 0
        for warning in results['warnings']:
            if 'split into' in str(warning).lower() or 'split sheet' in str(warning).lower():
                split_count += 1
        
        if split_count > 0:
            summary_parts.append(f"New tables created from splits: {split_count}")
    
    # Duplicate column names fixed
    total_dup_cols = 0
    if not results['meta_df'].empty:
        for _, row in results['meta_df'].iterrows():
            dup_cols = row.get('Duplicate column names fixed', 0)
            if pd.notna(dup_cols):
                total_dup_cols += int(dup_cols)
    
    if total_dup_cols > 0:
        summary_parts.append(f"Duplicate column names fixed: {total_dup_cols}")
    
    # Multi-row header detected
    max_header_depth = 1
    if not results['meta_df'].empty:
        for _, row in results['meta_df'].iterrows():
            depth = row.get('Header depth used', 1)
            if pd.notna(depth):
                max_header_depth = max(max_header_depth, int(depth))
    
    if max_header_depth > 1:
        summary_parts.append(f"Multi-row header detected: yes (max depth: {max_header_depth})")
    else:
        summary_parts.append("Multi-row header detected: no")
    
    # Totals flagged/removed
    total_flagged = 0
    total_removed = 0
    if not results['meta_df'].empty:
        for _, row in results['meta_df'].iterrows():
            flagged = row.get('Totals rows flagged', 0)
            removed = row.get('Totals rows dropped', 0)
            if pd.notna(flagged):
                total_flagged += int(flagged)
            if pd.notna(removed):
                total_removed += int(removed)
    
    if total_removed > 0:
        summary_parts.append(f"Totals flagged/removed: {total_flagged} flagged, {total_removed} removed")
    elif total_flagged > 0:
        summary_parts.append(f"Totals flagged: {total_flagged}")
    
    return summary_parts

# Helper function to format quality flag with badge HTML
def format_quality_flag_badge(flag_text):
    """Convert quality flag text to HTML badge."""
    flag_lower = str(flag_text).lower()
    
    if 'error' in flag_lower:
        badge_class = 'error'
        display_text = 'Error'
    elif 'warning' in flag_lower or 'review' in flag_lower:
        badge_class = 'warnings'
        display_text = 'Completed with warnings'
    elif 'new table' in flag_lower or 'new tables' in flag_lower:
        badge_class = 'completed'
        display_text = 'New tables created'
    elif 'success' in flag_lower or 'ok' in flag_lower or 'completed' in flag_lower:
        badge_class = 'completed'
        display_text = 'Completed'
    else:
        badge_class = 'completed'
        display_text = str(flag_text)
    
    return f'<span class="status-badge {badge_class}">{display_text}</span>'

# Helper function to recalculate quality flags with NEW TABLE status
def recalculate_quality_flags(meta_df):
    """Recalculate quality flags, showing NEW TABLE for split tables."""
    if meta_df.empty:
        return meta_df
    
    def new_flag(row):
        # ERROR if exception occurred
        if row["Clean status"] != "OK":
            return "ERROR"
        
        # Check for errors
        errors = str(row.get("Errors", "") if "Errors" in row.index else "")
        if errors and errors.strip() and errors.lower() != "nan":
            return "ERROR"
        
        # Check if this is a split table (NEW TABLE)
        warnings = str(row.get("Warnings", "") if "Warnings" in row.index else "")
        source_table_id = str(row.get("Source table ID", "") if "Source table ID" in row.index else "")
        
        # If it has a source_table_id and the warning mentions split, it's a new table
        is_split = False
        if source_table_id and source_table_id.strip() and source_table_id.lower() != "nan":
            if 'split into' in warnings.lower() or 'split sheet' in warnings.lower():
                is_split = True
        
        # Check warnings for split indicators
        if 'split into' in warnings.lower() or 'split sheet' in warnings.lower():
            is_split = True
        
        if is_split:
            # Only show NEW TABLES CREATED if no real warnings
            # Remove split warnings from check
            other_warnings = [w for w in warnings.split('; ') if w.strip() and 
                            'split into' not in w.lower() and 'split sheet' not in w.lower()]
            if not other_warnings:
                return "NEW TABLES CREATED"
        
        # Check for real warnings (non-split)
        has_real_warnings = False
        if warnings and warnings.strip() and warnings.lower() != "nan":
            other_warnings = [w for w in warnings.split('; ') if w.strip() and 
                            'split into' not in w.lower() and 'split sheet' not in w.lower()]
            if other_warnings:
                has_real_warnings = True
        
        if has_real_warnings:
            return "SUCCESS WITH WARNINGS"
        
        if row.get("Exact duplicate rows", 0) >= 100:
            return "SUCCESS WITH WARNINGS"
        
        if row.get("Totals rows dropped", 0) > 0:
            return "SUCCESS WITH WARNINGS"
        
        if row.get("Repeated header rows dropped", 0) > 0:
            return "SUCCESS WITH WARNINGS"
        
        # OK only if clean + no warnings
        return "SUCCESS"
    
    meta_df = meta_df.copy()
    meta_df["Quality flag"] = meta_df.apply(new_flag, axis=1)
    return meta_df

# Helper function to aggregate and classify messages
def aggregate_and_classify_messages(results):
    """Aggregate repeated warnings and classify messages into INFO, WARNINGS, and ERRORS."""
    info_list = []
    warnings_list = []
    errors_list = results['errors'].copy() if results['errors'] else []
    
    # Collect all messages by type
    split_info = []
    duplicate_col_warnings = []
    multirow_header_warnings = []
    other_warnings = []
    
    # Process warnings
    for warning in results['warnings']:
        warning_str = str(warning)
        if 'split into' in warning_str.lower() or 'split sheet' in warning_str.lower():
            split_info.append(warning_str)
        elif 'duplicate column' in warning_str.lower():
            duplicate_col_warnings.append(warning_str)
        elif 'multi-row header' in warning_str.lower():
            multirow_header_warnings.append(warning_str)
        else:
            other_warnings.append(warning_str)
    
    # Add table-specific messages and aggregate
    if not results['meta_df'].empty:
        for _, row in results['meta_df'].iterrows():
            table_name = row['Tab name']
            
            if pd.notna(row.get('Warnings', '')) and str(row['Warnings']).strip():
                for w in str(row['Warnings']).split('; '):
                    if w.strip():
                        if 'split into' in w.lower() or 'split sheet' in w.lower():
                            split_info.append(f"{table_name}: {w}")
                        elif 'duplicate column' in w.lower():
                            duplicate_col_warnings.append((table_name, w))
                        elif 'multi-row header' in w.lower() or 'header depth' in w.lower():
                            multirow_header_warnings.append((table_name, w))
                        else:
                            other_warnings.append(f"{table_name}: {w}")
            
            if pd.notna(row.get('Errors', '')) and str(row['Errors']).strip():
                for e in str(row['Errors']).split('; '):
                    if e.strip():
                        errors_list.append(f"{table_name}: {e}")
    
    # Aggregate duplicate column warnings
    if duplicate_col_warnings:
        total_dup = 0
        table_count = 0
        table_details = []
        for item in duplicate_col_warnings:
            if isinstance(item, tuple):
                table_name, msg = item
            else:
                table_name = "N/A"
                msg = item
            # Extract number from message
            import re
            match = re.search(r'(\d+)', msg)
            if match:
                total_dup += int(match.group(1))
            table_details.append((table_name, msg))
            table_count = len(set([t[0] for t in table_details]))
        
        if total_dup > 0:
            warnings_list.append({
                'summary': f"Duplicate column names fixed: total {total_dup} across {table_count} table(s)",
                'details': table_details,
                'type': 'duplicate_columns'
            })
    
    # Aggregate multi-row header warnings
    if multirow_header_warnings:
        max_depth = 1
        table_count = 0
        table_details = []
        for item in multirow_header_warnings:
            if isinstance(item, tuple):
                table_name, msg = item
            else:
                table_name = "N/A"
                msg = item
            import re
            match = re.search(r'depth[=\s]*(\d+)', msg, re.IGNORECASE)
            if match:
                max_depth = max(max_depth, int(match.group(1)))
            table_details.append((table_name, msg))
            table_count = len(set([t[0] for t in table_details]))
        
        warnings_list.append({
            'summary': f"Multi-row headers detected in {table_count} table(s) (max depth = {max_depth})",
            'details': table_details,
            'type': 'multirow_header'
        })
    
    # Add other warnings as individual items
    for w in other_warnings:
        warnings_list.append({
            'summary': w,
            'details': [],
            'type': 'other'
        })
    
    # Add split info
    if split_info:
        info_list.append({
            'summary': f"Sheet split into multiple tables: {len(split_info)} split(s) detected",
            'details': split_info,
            'type': 'split'
        })
    
    return info_list, warnings_list, errors_list

# Page header with feedback and logout buttons
col_header1, col_header2, col_header3 = st.columns([3, 1, 1], gap="small")
with col_header1:
    st.title("Spreadsheet Normalizer")
with col_header2:
    if st.button("Give feedback", key="feedback_button_header", help="Open feedback form"):
        st.session_state.feedback_modal_open = not st.session_state.get('feedback_modal_open', False)
with col_header3:
    if st.button("🔒 Logout", key="logout_button", help="Logout and clean up session"):
        # Clean up all session files before logout
        cleanup_session_files()
        st.session_state.authenticated = False
        st.rerun()

st.markdown("---")

# Feedback modal (displayed when button is clicked)
if st.session_state.get('feedback_modal_open', False):
    with st.container():
        st.markdown("### Share your feedback")
        feedback_text = st.text_area(
            "Share feedback or issues",
            placeholder="Tell us about your experience or report any problems...",
            height=100,
            key="feedback_input_modal",
            label_visibility="visible"
        )
        
        col_submit, col_cancel, col_spacer = st.columns([1, 1, 3])
        with col_submit:
            if st.button("Submit", key="feedback_submit_modal"):
                if feedback_text.strip():
                    # Save feedback to local file
                    feedback_dir = Path("feedback")
                    feedback_dir.mkdir(exist_ok=True)
                    
                    feedback_data = {
                        'timestamp': datetime.now().isoformat(),
                        'job_id': st.session_state.job_id,
                        'feedback': feedback_text
                    }
                    
                    feedback_file = feedback_dir / f"feedback_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                    with open(feedback_file, 'w') as f:
                        json.dump(feedback_data, f, indent=2)
                    
                    st.session_state.feedback_modal_open = False
                    st.markdown("""
                    <div style="background-color: #D6F5F0; color: #0D3A32; padding: 1rem; border-radius: 12px; 
                                border-left: 4px solid #00FDCF; margin: 1rem 0;">
                        <strong>Thank you for your feedback!</strong>
                    </div>
                    """, unsafe_allow_html=True)
                    time.sleep(1.5)
                    st.rerun()
                else:
                    st.warning("Please enter some feedback before submitting.")
        with col_cancel:
            if st.button("Cancel", key="feedback_cancel_modal"):
                st.session_state.feedback_modal_open = False
                st.rerun()

# Explanation section - split into three separate cards
# Card 1: What is normalization?
st.markdown("""
<div class="info-card">
    <h3>What is normalization?</h3>
    <p style="margin-bottom: 0; line-height: 1.6;">
        Data normalization prepares spreadsheets for databases and analytics tools. It transforms messy Excel files 
        into clean, structured data ready for SQL queries, analytics platforms, and AI tools. This tool applies 
        first normal form (1NF) by ensuring each cell contains a single value and rows are unique.
    </p>
</div>
""", unsafe_allow_html=True)

# Card 2: What it does
st.markdown("""
<div class="info-card">
    <h3>What it does</h3>
    <ul style="margin-top: 0.5rem; margin-bottom: 0; line-height: 1.8;">
        <li>Detects and separates multiple tables per sheet when needed</li>
        <li>Flattens multi-row headers and expands merged cells</li>
        <li>Standardizes column names and exports SQL-ready output (Excel/CSV) plus a readable run report</li>
    </ul>
</div>
""", unsafe_allow_html=True)

# Card 3: Possible results
st.markdown("""
<div class="info-card">
    <h3>Possible results</h3>
    <div style="margin-top: 0.5rem; display: flex; flex-direction: column; gap: 0.75rem;">
        <span class="status-badge completed"><strong>Completed:</strong> All tables normalized successfully</span>
        <span class="status-badge warnings"><strong>Completed with warnings:</strong> Normalized successfully, but review warnings to ensure data quality. Warnings can usually be bypassed.</span>
        <span class="status-badge error"><strong>Error:</strong> Issues found that must be fixed before database upload</span>
    </div>
</div>
""", unsafe_allow_html=True)

# File upload section
st.markdown("### Upload spreadsheet")
uploaded_files = st.file_uploader(
    "Choose Excel file(s) (.xlsx only, up to 5 files)",
    type=['xlsx'],
    help="Upload 1-5 Excel files to normalize. Single file shows full results; multiple files use batch mode.",
    accept_multiple_files=True
)

# Limit to 5 files
if uploaded_files and len(uploaded_files) > 5:
    st.error("Please upload a maximum of 5 files.")
    uploaded_files = uploaded_files[:5]

# Detect new file uploads - compare with last processed files
# Only clear state when files actually change, not when processing
if uploaded_files and not st.session_state.processing:
    uploaded_files_list = uploaded_files if isinstance(uploaded_files, list) else [uploaded_files]
    current_file_names = sorted([f.name for f in uploaded_files_list])
    
    # Check if files have changed from what was last processed
    if st.session_state.last_processed_files is not None:
        last_file_names = sorted(st.session_state.last_processed_files if isinstance(st.session_state.last_processed_files, list) else [st.session_state.last_processed_files])
        if current_file_names != last_file_names:
            # New files detected - clear processing state (but only if not currently processing)
            st.session_state.processing_complete = False
            st.session_state.results = None
            st.session_state.batch_results = None
elif not uploaded_files:
    # No files uploaded - clear state if we had processed files before (but only if not processing)
    if st.session_state.last_processed_files is not None and not st.session_state.processing:
        st.session_state.processing_complete = False
        st.session_state.results = None
        st.session_state.batch_results = None

# Output format selection
output_format_code = None
if uploaded_files and len(uploaded_files) > 0:
    st.markdown("### Choose output format")
    output_format = st.radio(
        "Choose output format:",
        ["Excel only", "CSV only", "Both Excel and CSV"],
        horizontal=True,
        help="Excel includes cleaned sheets plus META and TYPE_ANALYSIS sheets. CSV combines all tables into one file."
    )
    
    # Map selection to format code
    format_map = {
        "Excel only": "1",
        "CSV only": "2",
        "Both Excel and CSV": "3"
    }
    output_format_code = format_map[output_format]
    
    # Normalize button - always allows new run, even on same file
    if st.button("Normalize", type="primary", use_container_width=True):
        # Check rate limit before processing
        allowed, message = check_rate_limit()
        if not allowed:
            st.error(message)
            st.stop()
        
        # Always allow new processing run - reset states
        st.session_state.processing = True
        st.session_state.processing_complete = False
        # Record request for rate limiting
        record_request()
        # Keep results visible until new processing completes
        st.rerun()

# Processing indicator and execution - persistent status message
processing_placeholder = st.empty()

# Priority order: Processing > Success > New file detected
# IMPORTANT: Processing state takes absolute priority - if processing is True, show processing message
if st.session_state.processing:
    with processing_placeholder.container():
        # Show processing message with gray colors (loading state)
        st.markdown("""
        <div style="background-color: #EEEEEE; color: #1B1B1B; padding: 1.5rem; border-radius: 12px; 
                    border-left: 4px solid #999999; margin: 1rem 0; font-family: 'Poppins', sans-serif;">
            <strong style="font-weight: 600; font-size: 1.1rem;">Processing...</strong>
            <p style="margin: 0.5rem 0 0 0; font-size: 1rem;">Please wait while we normalize your spreadsheet.</p>
        </div>
        """, unsafe_allow_html=True)
    
    # Actually process (this runs while showing the processing message)
    # Streamlit will show the processing message above, then execute this code
    # This block will execute regardless of file state - processing takes priority
    try:
        # Create temp job folder
        job_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        job_dir = Path("temp_jobs") / job_id
        job_dir.mkdir(parents=True, exist_ok=True)
        
        # Track job directory for cleanup
        if job_dir not in st.session_state.session_job_dirs:
            st.session_state.session_job_dirs.append(job_dir)
        
        uploaded_files_list = uploaded_files if isinstance(uploaded_files, list) else [uploaded_files] if uploaded_files else []
        
        if not uploaded_files_list:
            st.error("Please upload at least one file.")
            st.session_state.processing = False
            st.session_state.processing_complete = False
            st.rerun()
        
        # Single file mode
        if len(uploaded_files_list) == 1:
            uploaded_file = uploaded_files_list[0]
            file_bytes = uploaded_file.getbuffer()
            file_name = uploaded_file.name
            
            input_path = job_dir / file_name
            with open(input_path, "wb") as f:
                f.write(file_bytes)
            
            # Call normalization function
            results = normalize_spreadsheet(
                input_path=input_path,
                output_format=output_format_code,
                output_dir=job_dir
            )
            
            # Create run report
            report_path = create_run_report_excel(
                results, job_id, file_name, output_format, job_dir
            )
            results['report_path'] = report_path
            
            # Store results in session state
            st.session_state.results = results
            st.session_state.batch_results = None
            st.session_state.job_id = job_id
            st.session_state.current_file_name = file_name
            st.session_state.current_output_format = output_format
            # Track processed files
            st.session_state.last_processed_files = [file_name]
            
            # Schedule cleanup of input file after processing (keep outputs for download)
            # Delete original uploaded file immediately after processing
            if input_path.exists():
                secure_delete_file(input_path)
            
            # Add to run history
            run_entry = {
                'job_id': job_id,
                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                'file_name': file_name,
                'status': results['status'],
                'tables_count': len(results['meta_df']) if not results['meta_df'].empty else 0,
                'results': results,
                'output_format': output_format
            }
            st.session_state.run_history.insert(0, run_entry)
        
        # Batch mode (multiple files)
        else:
            batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")
            batch_dir = Path("temp_jobs") / f"batch_{batch_id}"
            batch_dir.mkdir(parents=True, exist_ok=True)
            
            # Track batch directory for cleanup
            if batch_dir not in st.session_state.session_job_dirs:
                st.session_state.session_job_dirs.append(batch_dir)
            
            batch_results = []
            batch_summary = {
                'total_files': len(uploaded_files_list),
                'success_count': 0,
                'warning_count': 0,
                'error_count': 0,
                'total_tables': 0
            }
            
            for uploaded_file in uploaded_files_list:
                try:
                    file_name = uploaded_file.name
                    file_job_dir = batch_dir / file_name.replace('.xlsx', '').replace('.xls', '')
                    file_job_dir.mkdir(parents=True, exist_ok=True)
                    
                    input_path = file_job_dir / file_name
                    with open(input_path, "wb") as f:
                        f.write(uploaded_file.getbuffer())
                    
                    file_results = normalize_spreadsheet(
                        input_path=input_path,
                        output_format=output_format_code,
                        output_dir=file_job_dir
                    )
                    
                    # Delete original uploaded file immediately after processing
                    if input_path.exists():
                        secure_delete_file(input_path)
                    
                    report_path = create_run_report_excel(
                        file_results, f"{batch_id}_{file_name}", file_name, output_format, file_job_dir
                    )
                    file_results['report_path'] = report_path
                    file_results['file_name'] = file_name
                    file_results['job_dir'] = file_job_dir
                    
                    batch_results.append(file_results)
                    batch_summary['total_tables'] += len(file_results['meta_df']) if not file_results['meta_df'].empty else 0
                    
                    if file_results['status'] == 'success':
                        batch_summary['success_count'] += 1
                    elif file_results['status'] == 'partial':
                        batch_summary['warning_count'] += 1
                    else:
                        batch_summary['error_count'] += 1
                        
                except Exception as e:
                    batch_summary['error_count'] += 1
                    batch_results.append({
                        'file_name': uploaded_file.name,
                        'status': 'error',
                        'error': str(e),
                        'job_dir': None
                    })
            
            # Create batch ZIP
            zip_path = batch_dir / f"batch_{batch_id}.zip"
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for result in batch_results:
                    if result.get('job_dir') and result['job_dir'].exists():
                        for file_path in result['job_dir'].iterdir():
                            if file_path.is_file() and (file_path.suffix == '.xlsx' or file_path.suffix == '.csv'):
                                zipf.write(file_path, file_path.name)
            
            # Create batch summary report
            batch_report_path = batch_dir / f"batch_report_{batch_id}.xlsx"
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "BATCH_SUMMARY"
            ws.append(["Batch Summary"])
            ws.append([])
            ws.append(["Batch ID", batch_id])
            ws.append(["Total Files", batch_summary['total_files']])
            ws.append(["Successful", batch_summary['success_count']])
            ws.append(["With Warnings", batch_summary['warning_count']])
            ws.append(["Errors", batch_summary['error_count']])
            ws.append(["Total Tables", batch_summary['total_tables']])
            ws.append([])
            ws.append(["File Name", "Status", "Tables", "Errors"])
            for result in batch_results:
                tables_count = len(result.get('meta_df', pd.DataFrame())) if not result.get('meta_df', pd.DataFrame()).empty else 0
                errors = '; '.join(result.get('errors', [])) if result.get('errors') else ''
                ws.append([result.get('file_name', 'Unknown'), result.get('status', 'unknown').upper(), tables_count, errors])
            wb.save(batch_report_path)
            
            # Store batch results
            st.session_state.batch_results = {
                'batch_id': batch_id,
                'batch_dir': batch_dir,
                'zip_path': zip_path,
                'batch_report_path': batch_report_path,
                'summary': batch_summary,
                'results': batch_results,
                'output_format': output_format
            }
            st.session_state.results = None
            # Track processed files
            st.session_state.last_processed_files = [f.name for f in uploaded_files_list]
        
        st.session_state.processing = False
        st.session_state.processing_complete = True
        
        # Schedule cleanup of old temp files (background, non-blocking)
        # Note: Output files remain available for download until session ends
        st.rerun()
        
    except Exception as e:
        st.session_state.processing = False
        st.session_state.processing_complete = False
        st.error(f"Error during normalization: {str(e)}")
        st.exception(e)

elif st.session_state.processing_complete and (st.session_state.results is not None or st.session_state.batch_results is not None):
    # Only show success if current files match processed files
    if uploaded_files:
        uploaded_files_list = uploaded_files if isinstance(uploaded_files, list) else [uploaded_files]
        current_file_names = sorted([f.name for f in uploaded_files_list])
    else:
        current_file_names = []
    
    last_processed = st.session_state.last_processed_files if st.session_state.last_processed_files else []
    last_processed_sorted = sorted(last_processed if isinstance(last_processed, list) else [last_processed])
    
    # Show success only if files match
    if current_file_names == last_processed_sorted:
        with processing_placeholder.container():
            # Show success message with brand colors (green/teal)
            st.markdown("""
            <div style="background-color: #D6F5F0; color: #0D3A32; padding: 1.5rem; border-radius: 12px; 
                        border-left: 4px solid #00FDCF; margin: 1rem 0; font-family: 'Poppins', sans-serif;">
                <strong style="font-weight: 600; font-size: 1.1rem;">Successful.</strong>
                <p style="margin: 0.5rem 0 0 0; font-size: 1rem;">Please scroll below to download.</p>
            </div>
            """, unsafe_allow_html=True)
    else:
        # Files don't match - show new file detected
        if uploaded_files:
            with processing_placeholder.container():
                st.markdown("""
                <div style="background-color: #FFFFFF; color: #1B1B1B; padding: 1.5rem; border-radius: 12px; 
                            border-left: 4px solid #00FDCF; margin: 1rem 0; font-family: 'Poppins', sans-serif;">
                    <strong style="font-weight: 600; font-size: 1.1rem;">New file detected</strong>
                    <p style="margin: 0.5rem 0 0 0; font-size: 1rem;">Please click "Normalize" to process your spreadsheet.</p>
                </div>
                """, unsafe_allow_html=True)
elif uploaded_files:
    # Files uploaded but not processed yet - show new file detected
    uploaded_files_list = uploaded_files if isinstance(uploaded_files, list) else [uploaded_files]
    current_file_names = sorted([f.name for f in uploaded_files_list])
    
    last_processed = st.session_state.last_processed_files if st.session_state.last_processed_files else []
    last_processed_sorted = sorted(last_processed if isinstance(last_processed, list) else [last_processed])
    
    # Show "new file detected" if files are different or nothing has been processed yet
    if not last_processed_sorted or current_file_names != last_processed_sorted:
        with processing_placeholder.container():
            st.markdown("""
            <div style="background-color: #FFFFFF; color: #1B1B1B; padding: 1.5rem; border-radius: 12px; 
                        border-left: 4px solid #00FDCF; margin: 1rem 0; font-family: 'Poppins', sans-serif;">
                <strong style="font-weight: 600; font-size: 1.1rem;">New file detected</strong>
                <p style="margin: 0.5rem 0 0 0; font-size: 1rem;">Please click "Normalize" to process your spreadsheet.</p>
            </div>
            """, unsafe_allow_html=True)

# Results dashboard
if st.session_state.results is not None:
    results = st.session_state.results
    st.markdown("---")
    
    # Show file name, run id, timestamp
    st.markdown(f"**File:** {st.session_state.current_file_name} | **Run ID:** {st.session_state.job_id} | **Time:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    st.markdown("## Results")
    
    # Status banner with badge containing text
    if results['status'] == 'success':
        st.markdown('<div class="status-success"><span class="status-badge completed"><strong>Completed:</strong> All tables normalized successfully</span></div>', unsafe_allow_html=True)
    elif results['status'] == 'partial':
        st.markdown('<div class="status-warning"><span class="status-badge warnings"><strong>Completed with warnings:</strong> Normalized with warnings</span></div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="status-error"><span class="status-badge error"><strong>Error:</strong> Failed - errors encountered</span></div>', unsafe_allow_html=True)
    
    # Core changes summary
    summary_parts = generate_run_summary(results, st.session_state.current_file_name, st.session_state.current_output_format)
    st.markdown("### Summary")
    summary_html = '<div class="summary-card">'
    for part in summary_parts:
        summary_html += f'<div class="summary-bullet">• {part}</div>'
    summary_html += '</div>'
    st.markdown(summary_html, unsafe_allow_html=True)
    
    # Table overview
    st.markdown("### Table overview")
    if not results['meta_df'].empty:
        # Recalculate quality flags with NEW TABLE status
        meta_df_updated = recalculate_quality_flags(results['meta_df'])
        # Create simplified table view
        display_df = meta_df_updated[['Tab name', 'Rows after clean', 'Columns after clean', 'Quality flag']].copy()
        display_df.columns = ['Table name', 'Rows', 'Columns', 'Status']
        
        # Format quality flag for display (simplify the text)
        def format_status(flag_text):
            flag_lower = str(flag_text).lower()
            if 'error' in flag_lower:
                return 'Error'
            elif 'warning' in flag_lower or 'review' in flag_lower or 'success with warnings' in flag_lower:
                return 'Completed with warnings'
            elif 'new table' in flag_lower or 'new tables' in flag_lower:
                return 'New tables created'
            elif 'success' in flag_lower or 'ok' in flag_lower or 'completed' in flag_lower:
                return 'Completed'
            else:
                return str(flag_text)
        
        display_df['Status'] = display_df['Status'].apply(format_status)
        
        # Ensure Rows and Columns are integers
        display_df['Rows'] = display_df['Rows'].astype(int)
        display_df['Columns'] = display_df['Columns'].astype(int)
        
        # Use Streamlit's native dataframe display with custom styling
        st.dataframe(
            display_df,
            use_container_width=True,
            hide_index=True
        )
    
    # Downloads section (moved up, right after Results)
    st.markdown("### Downloads")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        excel_exists = results.get('excel_output_path') and results['excel_output_path'].exists()
        if excel_exists:
            with open(results['excel_output_path'], 'rb') as f:
                st.download_button(
                    label="Download Excel",
                    data=f.read(),
                    file_name=results['excel_output_path'].name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    disabled=False,
                    key="dl_excel_main"
                )
        else:
            st.download_button(
                label="Download Excel",
                data=b"",
                file_name="",
                disabled=True,
                key="dl_excel_main_disabled"
            )
    
    with col2:
        csv_exists = results.get('csv_output_path') and results['csv_output_path'].exists()
        if csv_exists:
            with open(results['csv_output_path'], 'rb') as f:
                st.download_button(
                    label="Download CSV",
                    data=f.read(),
                    file_name=results['csv_output_path'].name,
                    mime="text/csv",
                    disabled=False,
                    key="dl_csv_main"
                )
        else:
            st.download_button(
                label="Download CSV",
                data=b"",
                file_name="",
                disabled=True,
                key="dl_csv_main_disabled"
            )
    
    with col3:
        report_exists = results.get('report_path') and results['report_path'].exists()
        if report_exists:
            with open(results['report_path'], 'rb') as f:
                st.download_button(
                    label="Download report",
                    data=f.read(),
                    file_name=results['report_path'].name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    disabled=False,
                    key="dl_report_main"
                )
        else:
            st.download_button(
                label="Download report",
                data=b"",
                file_name="",
                disabled=True,
                key="dl_report_main_disabled"
            )
    
    # Warnings, Errors, Info sections
    info_list, warnings_list, errors_list = aggregate_and_classify_messages(results)
    
    # Render Info section
    if info_list:
        info_content = []
        for item in info_list:
            if isinstance(item, dict):
                info_content.append(f"<div style='margin: 0.5rem 0;'>ℹ️ <strong>{item['summary']}</strong></div>")
                if item.get('details') and len(item['details']) > 0:
                    detail_items = []
                    for detail in item['details'][:10]:  # Limit to first 10
                        detail_items.append(f"<div style='margin: 0.25rem 0; margin-left: 1.5rem;'>• {detail}</div>")
                    details_html = "".join(detail_items)
                    info_content.append(f"<div style='margin-left: 1rem; margin-top: 0.5rem; margin-bottom: 0.5rem;'>{details_html}</div>")
            else:
                info_content.append(f"<div style='margin: 0.5rem 0;'>ℹ️ {item}</div>")
        custom_expander("Info", expanded=False, content_html="".join(info_content))
    
    # Render Warnings section
    if warnings_list:
        warnings_content = []
        for item in warnings_list:
            if isinstance(item, dict):
                warnings_content.append(f"<div style='margin: 0.5rem 0;'>⚠️ <strong>{item['summary']}</strong></div>")
                if item.get('details') and len(item['details']) > 0:
                    detail_items = []
                    for detail in item['details'][:10]:  # Limit to first 10
                        if isinstance(detail, tuple):
                            detail_items.append(f"<div style='margin: 0.25rem 0; margin-left: 1.5rem;'>• <strong>{detail[0]}:</strong> {detail[1]}</div>")
                        else:
                            detail_items.append(f"<div style='margin: 0.25rem 0; margin-left: 1.5rem;'>• {detail}</div>")
                    details_html = "".join(detail_items)
                    warnings_content.append(f"<div style='margin-left: 1rem; margin-top: 0.5rem; margin-bottom: 0.5rem;'>{details_html}</div>")
            else:
                warnings_content.append(f"<div style='margin: 0.5rem 0;'>⚠️ {item}</div>")
        custom_expander("Warnings", expanded=False, content_html="".join(warnings_content))
    
    # Render Errors section
    if errors_list:
        errors_content = []
        for error in errors_list[:10]:
            errors_content.append(f"<div style='margin: 0.5rem 0;'>❌ {error}</div>")
        if len(errors_list) > 10:
            more_errors_html = "".join([f"<div style='margin: 0.5rem 0;'>❌ {error}</div>" for error in errors_list[10:]])
            errors_content.append(f"<div style='margin-top: 1rem;'><strong>Additional errors:</strong></div>{more_errors_html}")
        custom_expander("Errors", expanded=False, content_html="".join(errors_content))

# Batch results section
if st.session_state.batch_results is not None:
    batch = st.session_state.batch_results
    st.markdown("---")
    st.markdown(f"**Batch ID:** {batch['batch_id']} | **Time:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    st.markdown("## Batch Results")
    
    # Batch summary
    summary = batch['summary']
    if summary['error_count'] == 0 and summary['warning_count'] == 0:
        st.markdown('<div class="status-success"><strong>✓ All files normalized successfully</strong></div>', unsafe_allow_html=True)
    elif summary['error_count'] == 0:
        st.markdown('<div class="status-warning"><strong>⚠ Batch completed with warnings</strong></div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="status-error"><strong>✗ Batch completed with errors</strong></div>', unsafe_allow_html=True)
    
    # Batch summary stats
    st.markdown("### Batch Summary")
    summary_html = '<div class="summary-card">'
    summary_html += f'<div class="summary-bullet">• Total files: {summary["total_files"]}</div>'
    summary_html += f'<div class="summary-bullet">• Successful: {summary["success_count"]}</div>'
    summary_html += f'<div class="summary-bullet">• With warnings: {summary["warning_count"]}</div>'
    summary_html += f'<div class="summary-bullet">• Errors: {summary["error_count"]}</div>'
    summary_html += f'<div class="summary-bullet">• Total tables: {summary["total_tables"]}</div>'
    summary_html += '</div>'
    st.markdown(summary_html, unsafe_allow_html=True)
    
    # Batch downloads
    st.markdown("### Downloads")
    col1, col2 = st.columns(2)
    
    with col1:
        if batch['zip_path'].exists():
            with open(batch['zip_path'], 'rb') as f:
                st.download_button(
                    label="Download ZIP",
                    data=f.read(),
                    file_name=batch['zip_path'].name,
                    mime="application/zip",
                    key="dl_batch_zip"
                )
    
    with col2:
        if batch['batch_report_path'].exists():
            with open(batch['batch_report_path'], 'rb') as f:
                st.download_button(
                    label="Download batch report",
                    data=f.read(),
                    file_name=batch['batch_report_path'].name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_batch_report"
                )

# Run history section
if st.session_state.run_history:
    st.markdown("---")
    st.markdown("## Run history")
    st.markdown("_Session-only history. Previous runs from this session are shown below._")
    
    for idx, run in enumerate(st.session_state.run_history[:10]):  # Show last 10 runs
        status_class = run['status']
        if status_class == 'success':
            border_class = 'success'
        elif status_class == 'partial':
            border_class = 'warning'
        else:
            border_class = 'error'
        
        with st.container():
            st.markdown(f"""
            <div class="run-history-item {border_class}">
                <strong>{run['file_name']}</strong> | Run ID: {run['job_id']} | {run['timestamp']}<br>
                Status: {run['status'].upper()} | Tables: {run['tables_count']}
            </div>
            """, unsafe_allow_html=True)
            
            # Download buttons for this run
            col1, col2, col3 = st.columns(3)
            run_results = run['results']
            
            excel_exists = run_results.get('excel_output_path') and run_results['excel_output_path'].exists()
            csv_exists = run_results.get('csv_output_path') and run_results['csv_output_path'].exists()
            report_exists = run_results.get('report_path') and run_results['report_path'].exists()
            
            with col1:
                if excel_exists:
                    with open(run_results['excel_output_path'], 'rb') as f:
                        st.download_button(
                            label="Download Excel",
                            data=f.read(),
                            file_name=run_results['excel_output_path'].name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            disabled=False,
                            key=f"excel_{run['job_id']}"
                        )
                else:
                    st.download_button(
                        label="Download Excel",
                        data=b"",
                        file_name="",
                        disabled=True,
                        key=f"excel_disabled_{run['job_id']}"
                    )
            
            with col2:
                if csv_exists:
                    with open(run_results['csv_output_path'], 'rb') as f:
                        st.download_button(
                            label="Download CSV",
                            data=f.read(),
                            file_name=run_results['csv_output_path'].name,
                            mime="text/csv",
                            disabled=False,
                            key=f"csv_{run['job_id']}"
                        )
                else:
                    st.download_button(
                        label="Download CSV",
                        data=b"",
                        file_name="",
                        disabled=True,
                        key=f"csv_disabled_{run['job_id']}"
                    )
            
            with col3:
                if report_exists:
                    with open(run_results['report_path'], 'rb') as f:
                        st.download_button(
                            label="Download report",
                            data=f.read(),
                            file_name=run_results['report_path'].name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            disabled=False,
                            key=f"report_{run['job_id']}"
                        )
                else:
                    st.download_button(
                        label="Download report",
                        data=b"",
                        file_name="",
                        disabled=True,
                        key=f"report_disabled_{run['job_id']}"
                    )
        
        st.markdown("<br>", unsafe_allow_html=True)


# Stage 2 placeholders
st.markdown("---")
st.markdown("## Database connections (Stage 2)")
st.markdown("_Coming next - database integration features_")

col1, col2, col3 = st.columns(3)

with col1:
    st.button("Connect to Snowflake", disabled=True, use_container_width=True, help="Stage 2 - coming next")

with col2:
    st.button("Connect to Supabase", disabled=True, use_container_width=True, help="Stage 2 - coming next")

with col3:
    st.button("Connect to another SQL database", disabled=True, use_container_width=True, help="Stage 2 - coming next")

# CRITICAL FIX: Inject expander arrow fix at the end - this ensures it loads after all content
components.html("""
<style>
/* COMPLETE HIDE: All Material Icons elements */
.streamlit-expanderHeader i,
.streamlit-expanderHeader i[class*="material"],
.streamlit-expanderHeader [class*="material-icons"],
.streamlit-expanderHeader [class*="MaterialIcons"],
.streamlit-expanderHeader [class*="material"],
.streamlit-expanderHeader svg,
.streamlit-expanderHeader path,
.streamlit-expanderHeader g {
    display: none !important;
    visibility: hidden !important;
    width: 0 !important;
    height: 0 !important;
    opacity: 0 !important;
    font-size: 0 !important;
    position: absolute !important;
    left: -9999px !important;
    clip: rect(0,0,0,0) !important;
}

/* Header setup */
.streamlit-expanderHeader {
    position: relative !important;
    padding-right: 2rem !important;
    overflow: hidden !important;
}

/* Custom arrow on RIGHT using ::after */
.streamlit-expanderHeader::after {
    content: '▼' !important;
    position: absolute !important;
    right: 0.75rem !important;
    top: 50% !important;
    transform: translateY(-50%) !important;
    font-size: 0.85rem !important;
    color: #1B1B1B !important;
    z-index: 1000 !important;
    pointer-events: none !important;
    font-family: 'Poppins', sans-serif !important;
}

.streamlit-expanderHeader[aria-expanded="true"]::after {
    content: '▲' !important;
}
</style>
<script>
(function() {
    function fixExpanders() {
        const headers = document.querySelectorAll('.streamlit-expanderHeader');
        headers.forEach(header => {
            // Remove ALL child elements that might contain Material Icons
            Array.from(header.children).forEach(child => {
                const tag = child.tagName.toLowerCase();
                const classes = (child.className || '').toString().toLowerCase();
                const text = (child.textContent || '').trim().toLowerCase();
                
                if (tag === 'i' || 
                    classes.includes('material') || 
                    text.includes('keyboard_arrow') ||
                    text.includes('expand_more') ||
                    text.includes('expand_less') ||
                    text.includes('chevron_')) {
                    child.remove();
                }
            });
            
            // Remove SVG
            header.querySelectorAll('svg').forEach(s => s.remove());
            
            // Remove text nodes containing Material Icons names
            const walker = document.createTreeWalker(header, NodeFilter.SHOW_TEXT);
            const toRemove = [];
            let node;
            while (node = walker.nextNode()) {
                const text = node.textContent.trim();
                const lower = text.toLowerCase();
                if ((lower === 'keyboard_arrow_right' || 
                     lower === 'keyboard_arrow_down' ||
                     lower === 'keyboard_arrow_up' ||
                     lower.includes('keyboard_arrow') ||
                     lower.includes('expand_more') ||
                     lower.includes('expand_less') ||
                     lower.includes('chevron_')) && text.length < 30) {
                    toRemove.push(node);
                }
            }
            toRemove.forEach(n => {
                const parent = n.parentElement;
                if (parent && parent !== header) {
                    const parentText = parent.textContent.trim();
                    if (parentText === n.textContent.trim()) {
                        parent.remove();
                    } else {
                        n.remove();
                    }
                } else {
                    n.remove();
                }
            });
            
            // Create/update arrow span
            let arrow = header.querySelector('.exp-arrow-fix');
            if (!arrow) {
                arrow = document.createElement('span');
                arrow.className = 'exp-arrow-fix';
                header.style.position = 'relative';
                header.appendChild(arrow);
            }
            arrow.style.cssText = 'position:absolute!important;right:0.75rem!important;top:50%!important;transform:translateY(-50%)!important;font-size:0.85rem!important;color:#1B1B1B!important;z-index:1000!important;pointer-events:none!important;font-family:"Poppins",sans-serif!important;';
            arrow.textContent = header.getAttribute('aria-expanded') === 'true' ? '▲' : '▼';
        });
    }
    
    // Run immediately and repeatedly
    fixExpanders();
    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', fixExpanders);
    }
    
    // Run with intervals
    [10, 50, 100, 200, 500, 1000].forEach(d => setTimeout(fixExpanders, d));
    setInterval(fixExpanders, 200);
    
    // Watch for changes
    const obs = new MutationObserver(() => {
        setTimeout(fixExpanders, 10);
    });
    obs.observe(document.body, {
        childList: true, 
        subtree: true, 
        attributes: true, 
        attributeFilter: ['aria-expanded', 'class']
    });
    
    // Update on clicks
    document.addEventListener('click', (e) => {
        if (e.target.closest('.streamlit-expanderHeader')) {
            setTimeout(fixExpanders, 5);
        }
    }, true);
})();
</script>
""", height=0)

# Animated Data Particles - Side decorations
# Using a more direct approach that works with Streamlit's iframe structure
components.html("""
<script>
(function() {
    function getTargetDocument() {
        try {
            // Try to access parent window (main Streamlit document)
            if (window.parent !== window && window.parent.document) {
                return window.parent.document;
            }
        } catch(e) {
            // Cross-origin restriction - fall back to current document
            console.log('Cannot access parent, using current document');
        }
        return document;
    }
    
    const targetDoc = getTargetDocument();
    const targetBody = targetDoc ? targetDoc.body : null;
    
    if (!targetBody) {
        console.log('Cannot access target body');
        return;
    }
    
    // Inject CSS into the target document
    function injectStyles() {
        if (!targetDoc.getElementById('data-particles-styles')) {
            const style = targetDoc.createElement('style');
            style.id = 'data-particles-styles';
            style.textContent = `
                .data-particles-container {
                    position: fixed !important;
                    top: 0 !important;
                    left: 0 !important;
                    right: 0 !important;
                    bottom: 0 !important;
                    pointer-events: none !important;
                    z-index: 999999 !important;
                    overflow: hidden !important;
                }
                
                .data-particle {
                    position: absolute !important;
                    width: 8px !important;
                    height: 8px !important;
                    background-color: #00FDCF !important;
                    border-radius: 50% !important;
                    opacity: 0 !important;
                    pointer-events: none !important;
                    box-shadow: 0 0 8px rgba(0, 253, 207, 0.8) !important;
                }
                
                .data-particle.left {
                    left: 5% !important;
                    animation: particleFloatLeft 8s infinite ease-in-out !important;
                }
                
                .data-particle.right {
                    right: 5% !important;
                    animation: particleFloatRight 8s infinite ease-in-out !important;
                }
                
                @keyframes particleFloatLeft {
                    0% {
                        opacity: 0;
                        transform: translateY(100vh) translateX(-20px) scale(0.5);
                    }
                    10% {
                        opacity: 0.8;
                    }
                    50% {
                        opacity: 1;
                        transform: translateY(50vh) translateX(0px) scale(1);
                    }
                    90% {
                        opacity: 0.8;
                    }
                    100% {
                        opacity: 0;
                        transform: translateY(-10vh) translateX(20px) scale(0.5);
                    }
                }
                
                @keyframes particleFloatRight {
                    0% {
                        opacity: 0;
                        transform: translateY(100vh) translateX(20px) scale(0.5);
                    }
                    10% {
                        opacity: 0.8;
                    }
                    50% {
                        opacity: 1;
                        transform: translateY(50vh) translateX(0px) scale(1);
                    }
                    90% {
                        opacity: 0.8;
                    }
                    100% {
                        opacity: 0;
                        transform: translateY(-10vh) translateX(-20px) scale(0.5);
                    }
                }
                
                @media (max-width: 768px) {
                    .data-particles-container {
                        display: none !important;
                    }
                }
            `;
            if (targetDoc.head) {
                targetDoc.head.appendChild(style);
            } else {
                targetBody.appendChild(style);
            }
        }
    }
    
    function createParticleSystem() {
        try {
            if (!targetBody) return;
            
            // Inject styles first
            injectStyles();
            
            // Remove existing container if it exists
            const existing = targetDoc.getElementById('data-particles-container');
            if (existing) {
                existing.remove();
            }
            
            // Create container
            const container = targetDoc.createElement('div');
            container.id = 'data-particles-container';
            container.className = 'data-particles-container';
            targetBody.appendChild(container);
            
            // Number of particles
            const particleCount = 15;
            
            // Create particles with staggered delays
            for (let i = 0; i < particleCount; i++) {
                // Left side particles
                const leftParticle = targetDoc.createElement('div');
                leftParticle.className = 'data-particle left';
                const leftDelay = i * 0.6;
                const leftDuration = 7 + Math.random() * 5;
                leftParticle.style.animationDelay = leftDelay + 's';
                leftParticle.style.animationDuration = leftDuration + 's';
                const leftOpacity = 0.7 + Math.random() * 0.3;
                leftParticle.style.opacity = leftOpacity;
                leftParticle.style.backgroundColor = 'rgba(0, 253, 207, ' + leftOpacity + ')';
                leftParticle.style.left = (3 + Math.random() * 4) + '%';
                container.appendChild(leftParticle);
                
                // Right side particles
                const rightParticle = targetDoc.createElement('div');
                rightParticle.className = 'data-particle right';
                const rightDelay = i * 0.6 + 0.3;
                const rightDuration = 7 + Math.random() * 5;
                rightParticle.style.animationDelay = rightDelay + 's';
                rightParticle.style.animationDuration = rightDuration + 's';
                const rightOpacity = 0.7 + Math.random() * 0.3;
                rightParticle.style.opacity = rightOpacity;
                rightParticle.style.backgroundColor = 'rgba(0, 253, 207, ' + rightOpacity + ')';
                rightParticle.style.right = (3 + Math.random() * 4) + '%';
                container.appendChild(rightParticle);
            }
        } catch(e) {
            console.log('Particle system error:', e);
        }
    }
    
    // Initialize with multiple attempts
    function initParticles() {
        if (targetDoc && targetBody) {
            if (targetDoc.readyState === 'loading') {
                targetDoc.addEventListener('DOMContentLoaded', createParticleSystem);
            } else {
                createParticleSystem();
            }
        }
    }
    
    // Try immediately
    initParticles();
    
    // Recreate on delays to catch Streamlit reruns
    setTimeout(createParticleSystem, 100);
    setTimeout(createParticleSystem, 500);
    setTimeout(createParticleSystem, 1000);
    setTimeout(createParticleSystem, 2000);
    setTimeout(createParticleSystem, 3000);
    
    // Watch for changes
    if (targetBody) {
        const observer = new MutationObserver(function() {
            const container = targetDoc.getElementById('data-particles-container');
            if (!container) {
                setTimeout(createParticleSystem, 100);
            }
        });
        
        observer.observe(targetBody, {
            childList: true,
            subtree: false
        });
    }
})();
</script>
""", height=0)

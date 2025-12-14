import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from typing import Tuple, Dict, List, Optional, Any
from datetime import datetime
import uuid
import shutil
import csv
import re
import math


# -----------------------------
# FILE SELECTION
# -----------------------------

def choose_input_file() -> Path:
    """
    Detect Excel files in the folder (.xlsx and .xls).
    If multiple exist, ask user to choose.
    Ignore temporary / cleaned files.
    """
    # Look for both .xlsx and .xls files
    xlsx_files = [
        f for f in Path(".").glob("*.xlsx")
        if not f.name.startswith("_") and not f.name.startswith("clean_")
    ]
    xls_files = [
        f for f in Path(".").glob("*.xls")
        if not f.name.startswith("_") and not f.name.startswith("clean_")
        and not f.name.endswith(".xlsx")  # Avoid duplicates
    ]
    candidates = xlsx_files + xls_files

    if not candidates:
        raise FileNotFoundError(
            "No Excel files (.xlsx or .xls) found. Place an Excel file here and run again."
        )

    if len(candidates) == 1:
        return candidates[0]

    print("\nMultiple Excel files found:")
    for idx, f in enumerate(candidates, 1):
        print(f" {idx}. {f.name}")

    while True:
        choice = input("Select file number: ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(candidates):
            return candidates[int(choice) - 1]
        print("Invalid selection. Try again.")


# -----------------------------
# OUTPUT FORMAT SELECTION
# -----------------------------

def choose_output_format() -> str:
    print("\nChoose output format:")
    print(" 1 = Excel (.xlsx)")
    print(" 2 = CSV (single file, all tabs stacked)")
    print(" 3 = Both Excel + CSV")

    while True:
        choice = input("Enter 1 / 2 / 3: ").strip()
        if choice in ("1", "2", "3"):
            return choice
        print("Invalid selection.")


# -----------------------------
# MERGED CELL EXPANSION
# -----------------------------

def expand_merged_cells(input_path: Path, output_path: Path):
    """
    Copy value from top-left of each merged range into all cells, then unmerge.
    Only works with .xlsx files (openpyxl limitation).
    For .xls files, merged cells are not expanded.
    """
    # openpyxl only supports .xlsx format
    if input_path.suffix.lower() == '.xls':
        # For .xls files, just copy the file (can't expand merged cells with openpyxl)
        import shutil
        shutil.copy2(input_path, output_path)
        return
    
    wb = load_workbook(input_path)

    for ws in wb.worksheets:
        for merged_range in list(ws.merged_cells.ranges):
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            min_col = merged_range.min_col
            max_col = merged_range.max_col

            top_left_value = ws.cell(row=min_row, column=min_col).value
            ws.unmerge_cells(str(merged_range))

            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    ws.cell(row=r, column=c, value=top_left_value)

    wb.save(output_path)


# -----------------------------
# SQL RESERVED WORDS (Snowflake)
# -----------------------------

# Snowflake SQL reserved words that cannot be used as unquoted identifiers
SQL_RESERVED_WORDS = {
    'all', 'alter', 'and', 'any', 'as', 'asc', 'between', 'by', 'case', 'cast',
    'check', 'column', 'connect', 'constraint', 'create', 'cross', 'current',
    'current_date', 'current_time', 'current_timestamp', 'database', 'delete',
    'desc', 'distinct', 'drop', 'else', 'end', 'exists', 'false', 'following',
    'for', 'from', 'full', 'grant', 'group', 'having', 'in', 'increment',
    'inner', 'insert', 'intersect', 'into', 'is', 'join', 'lateral', 'left',
    'like', 'localtime', 'localtimestamp', 'minus', 'natural', 'not', 'null',
    'of', 'on', 'or', 'order', 'organization', 'outer', 'over', 'partition',
    'preceding', 'primary', 'range', 'references', 'revoke', 'right', 'rlike',
    'row', 'rows', 'sample', 'select', 'set', 'some', 'start', 'table', 'tablesample',
    'then', 'to', 'trigger', 'true', 'union', 'unique', 'update', 'using',
    'values', 'view', 'when', 'whenever', 'where', 'with'
}


# -----------------------------
# HEADER DETECTION + CLEANING
# -----------------------------

def detect_header_row(df: pd.DataFrame) -> int:
    """
    Detect the header row as the first row with enough non-empty cells.
    (Legacy function - kept for backward compatibility)
    """
    df_tmp = df.replace("", pd.NA)
    non_empty_counts = df_tmp.notna().sum(axis=1)
    min_non_empty = max(2, int(df_tmp.shape[1] * 0.3))

    for idx, count in non_empty_counts.items():
        if count >= min_non_empty:
            return idx

    return df.index[0]


def detect_header_depth(df: pd.DataFrame) -> Tuple[int, int]:
    """
    Detect header depth (1-3 rows) by analyzing row patterns.
    Looks for consecutive rows with high non-empty density.
    
    Returns:
        tuple: (start_row_index, depth)
    """
    df_tmp = df.replace("", pd.NA)
    non_empty_counts = df_tmp.notna().sum(axis=1)
    min_non_empty = max(2, int(df_tmp.shape[1] * 0.3))
    
    # Find first row with enough non-empty cells (header start)
    start_idx = None
    for idx, count in non_empty_counts.items():
        if count >= min_non_empty:
            start_idx = idx
            break
    
    if start_idx is None:
        return df.index[0], 1
    
    # Check if next 1-2 rows also look like headers
    # Criteria: high non-empty density AND low numeric content (headers are usually text)
    depth = 1
    max_depth = min(3, len(df_tmp) - start_idx)
    
    for check_depth in range(2, max_depth + 1):
        check_idx = start_idx + check_depth - 1
        if check_idx not in df_tmp.index:
            break
        
        row = df_tmp.loc[check_idx]
        non_empty_count = row.notna().sum()
        non_empty_ratio = non_empty_count / len(row) if len(row) > 0 else 0
        
        # Check if this row has enough non-empty cells to be a header row
        if non_empty_count >= min_non_empty and non_empty_ratio >= 0.3:
            # Check if it's mostly non-numeric (typical of headers)
            numeric_count = 0
            for val in row:
                # Skip if val is a Series (shouldn't happen, but defensive)
                if isinstance(val, pd.Series):
                    continue
                # Use safe helper functions to avoid Series boolean ambiguity
                if safe_notna(val):
                    try:
                        val_str = safe_str_strip(val)
                        float(val_str.replace(',', '').replace('$', '').replace('%', ''))
                        numeric_count += 1
                    except (ValueError, AttributeError):
                        pass
            
            numeric_ratio = numeric_count / non_empty_count if non_empty_count > 0 else 1.0
            
            # If row has good density and is mostly non-numeric, it's likely a header row
            if numeric_ratio < 0.5:  # Less than 50% numeric suggests header
                depth = check_depth
            else:
                break
        else:
            break
    
    return start_idx, depth


def flatten_multirow_headers(df: pd.DataFrame, start_idx: int, depth: int) -> pd.Series:
    """
    Flatten multiple header rows into a single header row.
    Joins non-empty header parts per column with underscores.
    
    Args:
        df: Raw dataframe
        start_idx: Starting row index of headers
        depth: Number of header rows (1-3)
    
    Returns:
        Series: Flattened header row
    """
    if depth == 1:
        return df.loc[start_idx]
    
    # Get header rows
    header_rows = []
    for i in range(depth):
        row_idx = start_idx + i
        if row_idx in df.index:
            header_rows.append(df.loc[row_idx])
    
    if not header_rows:
        return df.loc[start_idx]
    
    # Build flattened headers column by column
    num_cols = len(header_rows[0])
    flattened = []
    
    for col_idx in range(num_cols):
        parts = []
        for row in header_rows:
            val = row.iloc[col_idx] if col_idx < len(row) else None
            # Use safe helper functions to avoid Series boolean ambiguity
            if val is not None and not isinstance(val, pd.Series):
                if safe_notna(val):
                    val_str = safe_str_strip(val)
                    if val_str:
                        parts.append(val_str)
        
        if parts:
            # Join parts with underscore, but avoid duplicate consecutive parts
            cleaned_parts = []
            prev = None
            for part in parts:
                if part != prev:
                    cleaned_parts.append(part)
                prev = part
            flattened_val = '_'.join(cleaned_parts)
        else:
            flattened_val = f"unnamed_col_{col_idx + 1}"
        
        flattened.append(flattened_val)
    
    # Return as Series with simple integer index (will be used as column names)
    return pd.Series(flattened)


def detect_table_regions(df: pd.DataFrame) -> List[Dict]:
    """
    Detect distinct table regions within a sheet.
    Tables can be separated by blank columns or blank rows.
    Uses a simple approach: find completely blank separator rows/columns.
    
    Args:
        df: Raw dataframe (before header detection)
    
    Returns:
        list: List of dicts with keys: {
            'min_row': int,
            'max_row': int,
            'min_col': int,
            'max_col': int,
            'bounds': str  # e.g., "R1C1:R10C5"
        }
    """
    if df.empty:
        return []
    
    # Convert to string and normalize blanks
    df_str = df.astype(str)
    df_str = df_str.replace(['nan', 'None', 'null', ''], pd.NA)
    
    # Find completely empty rows and columns (separators)
    empty_rows_mask = df_str.isna().all(axis=1)
    empty_cols_mask = df_str.isna().all(axis=0)
    
    # Get row and column positions
    row_positions = list(range(len(df)))
    col_positions = list(range(len(df.columns)))
    
    # Find separator rows (completely blank rows)
    separator_rows = [i for i, empty in enumerate(empty_rows_mask) if empty]
    
    # Find separator columns (completely blank columns)
    separator_cols = [i for i, empty in enumerate(empty_cols_mask) if empty]
    
    # If no separators, return single region
    if not separator_rows and not separator_cols:
        return [{
            'min_row': df.index[0],
            'max_row': df.index[-1],
            'min_col': df.columns[0],
            'max_col': df.columns[-1],
            'bounds': f"R1C1:R{len(df)}C{len(df.columns)}"
        }]
    
    regions = []
    
    # Split by rows first
    row_ranges = []
    start = 0
    for sep_row in separator_rows:
        if start < sep_row:
            row_ranges.append((start, sep_row - 1))
        start = sep_row + 1
    if start < len(df):
        row_ranges.append((start, len(df) - 1))
    
    # Split by columns
    col_ranges = []
    start = 0
    for sep_col in separator_cols:
        if start < sep_col:
            col_ranges.append((start, sep_col - 1))
        start = sep_col + 1
    if start < len(df.columns):
        col_ranges.append((start, len(df.columns) - 1))
    
    # Create regions from row and column ranges
    for row_start, row_end in row_ranges:
        for col_start, col_end in col_ranges:
            # Extract region
            region_df = df_str.iloc[row_start:row_end+1, col_start:col_end+1]
            
            # Check if region has sufficient data (at least 30% non-empty)
            non_empty_count = region_df.notna().sum().sum()
            total_count = region_df.size
            density = non_empty_count / total_count if total_count > 0 else 0
            
            # Also check minimum size (at least 2x2)
            if (row_end - row_start + 1) >= 2 and (col_end - col_start + 1) >= 2 and density >= 0.3:
                regions.append({
                    'min_row': df.index[row_start],
                    'max_row': df.index[row_end],
                    'min_col': df.columns[col_start],
                    'max_col': df.columns[col_end],
                    'bounds': f"R{row_start+1}C{col_start+1}:R{row_end+1}C{col_end+1}"
                })
    
    # If no valid regions found, return entire sheet as one region
    if not regions:
        return [{
            'min_row': df.index[0],
            'max_row': df.index[-1],
            'min_col': df.columns[0],
            'max_col': df.columns[-1],
            'bounds': f"R1C1:R{len(df)}C{len(df.columns)}"
        }]
    
    return regions


def normalise_headers(headers) -> Tuple[list, int]:
    """
    Standardise column names: lowercase, underscores, no special characters.
    Ensures all headers are SQL-compatible:
    - Cannot start with numbers (prefixes with 'col_')
    - Escapes SQL reserved words (adds '_col' suffix)
    - Validates length (truncates to 255 chars for Snowflake)
    - Ensures all headers are non-empty and unique.
    
    Returns:
        tuple: (normalized_headers_list, duplicate_count)
    """
    # Step 1: Normalize (lowercase, underscores, remove special chars)
    clean = (
        pd.Series(headers)
        .astype(str)
        .str.strip()
        .str.lower()
        .str.replace(r"\s+", "_", regex=True)
        .str.replace(r"[^\w]", "", regex=True)
    )
    
    # Step 2: Replace empty headers with "unnamed_col_N"
    unnamed_counter = 1
    clean_list = []
    for i, val in enumerate(clean):
        if val == "" or val == "nan":
            clean_list.append(f"unnamed_col_{unnamed_counter}")
            unnamed_counter += 1
        else:
            clean_list.append(val)
    
    # Step 3: Fix SQL-incompatible column names
    sql_safe_list = []
    for header in clean_list:
        # Check if starts with number (SQL doesn't allow this)
        if header and header[0].isdigit():
            header = f"col_{header}"
        
        # Check if it's a SQL reserved word
        if header in SQL_RESERVED_WORDS:
            header = f"{header}_col"
        
        # Validate length (Snowflake max identifier length is 255)
        if len(header) > 255:
            header = header[:252] + "..."
        
        sql_safe_list.append(header)
    
    # Step 4: Ensure uniqueness by adding suffixes (_2, _3, etc.)
    seen = {}
    final_headers = []
    duplicate_count = 0
    
    for header in sql_safe_list:
        if header in seen:
            seen[header] += 1
            new_header = f"{header}_{seen[header]}"
            # Re-check length after adding suffix
            if len(new_header) > 255:
                new_header = new_header[:252] + "..."
            final_headers.append(new_header)
            duplicate_count += 1
        else:
            seen[header] = 1
            final_headers.append(header)
    
    return final_headers, duplicate_count


def detect_repeated_headers(df_data: pd.DataFrame, normalized_headers: List[str]) -> Tuple[pd.DataFrame, int]:
    """
    Detect and remove rows that match the header pattern (repeated headers in data).
    
    Args:
        df_data: DataFrame with normalized headers already set
        normalized_headers: List of normalized header names
    
    Returns:
        tuple: (cleaned_dataframe, count_of_removed_rows)
    """
    if df_data.empty or len(normalized_headers) == 0:
        return df_data, 0
    
    # Convert headers to a set of normalized strings for comparison
    header_set = set(str(h).strip().lower() for h in normalized_headers if str(h).strip())
    
    if not header_set:
        return df_data, 0
    
    # Check each row to see if it matches the header pattern
    rows_to_drop = []
    
    for idx in df_data.index:
        # Get row - handle case where duplicate columns might return DataFrame
        row_data = df_data.loc[idx]
        # If we got a DataFrame (duplicate columns), convert to Series
        if isinstance(row_data, pd.DataFrame):
            row_values = row_data.iloc[0].astype(str)
        else:
            row_values = pd.Series(row_data).astype(str)
        
        # Normalize row values for comparison
        # Ensure we iterate over scalar values, not Series
        row_set = set()
        for v in row_values:
            # Skip if v is a Series (shouldn't happen, but defensive)
            if isinstance(v, pd.Series):
                continue
            v_str = str(v).strip()
            if v_str:
                row_set.add(v_str.lower())
        
        # If the row contains a significant overlap with headers, it's likely a repeated header
        # We use a threshold: if >= 70% of non-empty row values match headers, consider it a header row
        if len(row_set) > 0:
            overlap = len(row_set.intersection(header_set))
            overlap_ratio = overlap / len(row_set) if len(row_set) > 0 else 0
            
            # Also check if the row has similar structure (many matches in same positions)
            position_matches = sum(
                1 for i, h in enumerate(normalized_headers)
                if i < len(row_values) and str(row_values.iloc[i]).strip().lower() == str(h).strip().lower()
            )
            position_match_ratio = position_matches / len(normalized_headers) if len(normalized_headers) > 0 else 0
            
            # If either overlap ratio or position match ratio is high, it's likely a header row
            if overlap_ratio >= 0.7 or (position_match_ratio >= 0.7 and len(row_set) >= len(header_set) * 0.5):
                rows_to_drop.append(idx)
    
    if rows_to_drop:
        df_cleaned = df_data.drop(index=rows_to_drop).reset_index(drop=True)
        return df_cleaned, len(rows_to_drop)
    
    return df_data, 0


def detect_total_rows(df_data: pd.DataFrame) -> Tuple[pd.DataFrame, int, int]:
    """
    Detect rows that contain totals, subtotals, or grand totals.
    Flags them but only removes if high confidence.
    
    Args:
        df_data: DataFrame with data rows
    
    Returns:
        tuple: (cleaned_dataframe, flagged_count, dropped_count)
    """
    if df_data.empty:
        return df_data, 0, 0
    
    # Keywords that indicate totals/subtotals (case-insensitive)
    total_keywords = [
        'total', 'subtotal', 'grand total', 'grand_total',
        'sum', 'summary', 'totals', 'subtotals',
        '合计', '总计', '小计'  # Chinese variants
    ]
    
    # Create flag column
    df_data = df_data.copy()
    df_data['__is_total_row'] = False
    
    flagged_indices = []
    dropped_indices = []
    
    for idx in df_data.index:
        # Get row - handle case where duplicate columns might return DataFrame
        row_data = df_data.loc[idx]
        # If we got a DataFrame (duplicate columns), convert to Series
        if isinstance(row_data, pd.DataFrame):
            row = row_data.iloc[0]
        else:
            row = pd.Series(row_data)
        
        # Build row string safely, ensuring we handle scalar values
        row_parts = []
        for v in row:
            # Skip if v is a Series (shouldn't happen, but defensive)
            if isinstance(v, pd.Series):
                continue
            if safe_notna(v):
                v_str = safe_str_strip(v)
                if v_str:
                    row_parts.append(v_str.lower())
        row_str = ' '.join(row_parts)
        
        # Check if row contains total keywords
        is_total = any(keyword in row_str for keyword in total_keywords)
        
        if is_total:
            flagged_indices.append(idx)
            df_data.loc[idx, '__is_total_row'] = True
            
            # High confidence indicators for dropping:
            # 1. Contains "grand total" or "subtotal" (more specific)
            # 2. Row has mostly numeric values (typical of summary rows)
            # 3. Row appears near the end of the dataframe (common location)
            
            has_specific_keyword = any(kw in row_str for kw in ['grand total', 'subtotal', 'grand_total'])
            
            # Check if row has mostly numeric values (indicator of summary)
            numeric_count = 0
            non_empty_count = 0
            for val in row:
                # Skip if val is a Series (shouldn't happen, but defensive)
                if isinstance(val, pd.Series):
                    continue
                # Use safe helper functions to avoid Series boolean ambiguity
                if safe_notna(val):
                    val_str = safe_str_strip(val)
                    if val_str:
                        non_empty_count += 1
                        try:
                            float(val_str.replace(',', '').replace('$', '').replace('%', ''))
                            numeric_count += 1
                        except (ValueError, AttributeError):
                            pass
            
            mostly_numeric = (numeric_count / non_empty_count >= 0.7) if non_empty_count > 0 else False
            
            # Check if near end (last 10% of rows)
            is_near_end = idx >= len(df_data) * 0.9
            
            # High confidence: has specific keyword AND (mostly numeric OR near end)
            high_confidence = has_specific_keyword and (mostly_numeric or is_near_end)
            
            if high_confidence:
                dropped_indices.append(idx)
    
    # Remove high-confidence total rows
    if dropped_indices:
        df_cleaned = df_data.drop(index=dropped_indices).reset_index(drop=True)
    else:
        df_cleaned = df_data
    
    return df_cleaned, len(flagged_indices), len(dropped_indices)


def detect_context_columns(df: pd.DataFrame, normalized_headers: List[str]) -> List[str]:
    """
    Detect columns that are likely context/category columns.
    Criteria:
    - Low cardinality (few unique values relative to row count)
    - High percentage of blanks (many empty cells)
    - Column name suggests hierarchy (contains common category words)
    
    Args:
        df: DataFrame with data
        normalized_headers: List of normalized header names
    
    Returns:
        list: Column names identified as context columns
    """
    if df.empty or len(normalized_headers) == 0:
        return []
    
    context_columns = []
    n_rows = len(df)
    
    # Keywords that suggest category/context columns
    category_keywords = ['category', 'group', 'section', 'region', 'department', 
                         'division', 'type', 'class', 'level', 'hierarchy']
    
    for col in df.columns:
        if col in ['__possible_duplicate', '__is_total_row']:
            continue
        
        # Ensure we get a Series (handle duplicate column names or edge cases)
        try:
            # Use iloc to get column by position if name access fails, or ensure Series
            if col in df.columns:
                col_data = df[col]
                # If we got a DataFrame (duplicate columns), take first column
                if isinstance(col_data, pd.DataFrame):
                    series = col_data.iloc[:, 0].astype(str)
                else:
                    series = pd.Series(col_data).astype(str)
            else:
                continue
        except (KeyError, IndexError):
            continue
        
        # Ensure series is actually a Series
        if not isinstance(series, pd.Series):
            continue
        
        # Count blanks/empty values - use safe helper functions
        blank_count = 0
        for v in series:
            # Use safe helper functions to avoid Series boolean ambiguity
            if safe_isna(v):
                blank_count += 1
                continue
            v_str = safe_str_strip(v)
            if v_str in ['', 'nan', 'None', 'null']:
                blank_count += 1
        blank_ratio = blank_count / n_rows if n_rows > 0 else 0
        
        # Count unique non-blank values - use safe string operations
        non_blank = pd.Series(dtype=object)  # Initialize to empty Series
        unique_count = 0
        cardinality_ratio = 1.0
        
        try:
            # Convert to string and filter safely
            series_str = series.astype(str)
            
            # Build mask step by step to avoid ambiguous boolean operations
            # Use explicit boolean Series operations
            mask_notna = series_str.str.strip().notna()
            mask_not_empty = (series_str.str.strip() != '')
            mask_not_nan = (series_str.str.strip() != 'nan')
            mask_not_none = (series_str.str.strip() != 'None')
            
            # Combine masks using & operator (element-wise for Series)
            # Check each mask is a Series before combining
            if (isinstance(mask_notna, pd.Series) and 
                isinstance(mask_not_empty, pd.Series) and
                isinstance(mask_not_nan, pd.Series) and
                isinstance(mask_not_none, pd.Series)):
                non_blank_mask = mask_notna & mask_not_empty & mask_not_nan & mask_not_none
                # Ensure mask is a Series before using for indexing
                if isinstance(non_blank_mask, pd.Series):
                    non_blank = series[non_blank_mask]
                else:
                    # Fallback: filter manually using list comprehension
                    # Use explicit checks to avoid Series boolean ambiguity
                    non_blank_list = []
                    for v in series:
                        # Skip if v is a Series (shouldn't happen, but defensive)
                        if isinstance(v, pd.Series):
                            continue
                        # Use explicit checks, not boolean operations on potential Series
                        if safe_notna(v):
                            v_str = safe_str_strip(v)
                            if v_str and v_str.lower() not in ['nan', 'none', 'null']:
                                non_blank_list.append(v)
                    non_blank = pd.Series(non_blank_list) if non_blank_list else pd.Series(dtype=object)
            else:
                # Fallback: filter manually using list comprehension
                # Use explicit checks to avoid Series boolean ambiguity
                non_blank_list = []
                for v in series:
                    # Skip if v is a Series (shouldn't happen, but defensive)
                    if isinstance(v, pd.Series):
                        continue
                    # Use safe helper functions to avoid Series boolean ambiguity
                    if safe_notna(v):
                        v_str = safe_str_strip(v)
                        if v_str and v_str.lower() not in ['nan', 'none', 'null']:
                            non_blank_list.append(v)
                non_blank = pd.Series(non_blank_list) if non_blank_list else pd.Series(dtype=object)
            
            # Use explicit length check to avoid Series boolean ambiguity
            non_blank_len = len(non_blank) if isinstance(non_blank, pd.Series) else 0
            unique_count = non_blank.nunique() if non_blank_len > 0 else 0
            cardinality_ratio = unique_count / non_blank_len if non_blank_len > 0 else 1.0
        except (AttributeError, TypeError, ValueError) as e:
            # Fallback if .str accessor fails or boolean operation fails
            try:
                # More defensive approach
                non_blank_list = []
                for val in series:
                    # Skip if val is a Series (shouldn't happen, but defensive)
                    if isinstance(val, pd.Series):
                        continue
                    # Use explicit checks to avoid Series boolean ambiguity
                    if safe_notna(val):
                        val_str = safe_str_strip(val)
                        if val_str and val_str.lower() not in ['nan', 'none', 'null']:
                            non_blank_list.append(val)
                non_blank = pd.Series(non_blank_list) if non_blank_list else pd.Series(dtype=object)
                unique_count = non_blank.nunique() if len(non_blank) > 0 else 0
                cardinality_ratio = unique_count / len(non_blank) if len(non_blank) > 0 else 1.0
            except Exception:
                # Ultimate fallback
                non_blank = pd.Series(dtype=object)
                unique_count = 0
                cardinality_ratio = 1.0
        
        # Check if column name suggests category
        col_lower = str(col).lower()
        has_category_keyword = any(kw in col_lower for kw in category_keywords)
        
        # Detect context column if:
        # 1. High blank ratio (>= 30%) AND low cardinality (< 20% unique), OR
        # 2. Very low cardinality (< 10% unique) regardless of blanks, OR
        # 3. Has category keyword AND (high blanks OR low cardinality)
        # Use explicit length check to avoid Series boolean ambiguity
        non_blank_len = len(non_blank) if isinstance(non_blank, pd.Series) else 0
        is_context = (
            (blank_ratio >= 0.3 and cardinality_ratio < 0.2) or
            (cardinality_ratio < 0.1 and non_blank_len > 0) or
            (has_category_keyword and (blank_ratio >= 0.2 or cardinality_ratio < 0.3))
        )
        
        if is_context:
            context_columns.append(col)
    
    return context_columns


def fill_down_context(df: pd.DataFrame, context_columns: List[str]) -> pd.DataFrame:
    """
    Fill down blank cells in context columns until next non-blank value.
    This handles the "Tetris" case where category labels only appear once.
    
    Args:
        df: DataFrame with data
        context_columns: List of column names to apply fill-down
    
    Returns:
        DataFrame with filled context columns
    """
    if df.empty or not context_columns:
        return df
    
    df_filled = df.copy()
    
    for col in context_columns:
        if col not in df_filled.columns:
            continue
        
        # Ensure we get a Series (handle duplicate column names or edge cases)
        try:
            col_data = df_filled[col]
            # If we got a DataFrame (duplicate columns), take first column
            if isinstance(col_data, pd.DataFrame):
                series = col_data.iloc[:, 0].astype(str)
            else:
                series = pd.Series(col_data).astype(str)
        except (KeyError, IndexError, TypeError):
            continue
        
        # Ensure series is actually a Series
        if not isinstance(series, pd.Series):
            continue
        
        filled_series = series.copy()
        
        # Track last non-blank value
        last_value = None
        
        for idx in df_filled.index:
            val = series.loc[idx]
            # Use explicit checks to avoid Series boolean ambiguity
            if isinstance(val, pd.Series):
                val_str = ''
            else:
                if safe_notna(val):
                    val_str = safe_str_strip(val)
                else:
                    val_str = ''
            
            # Check if blank/empty
            is_blank = (val_str == '' or val_str.lower() in ['nan', 'none', 'null'])
            
            if is_blank:
                # Fill with last non-blank value if available
                if last_value is not None:
                    filled_series.loc[idx] = last_value
            else:
                # Update last non-blank value
                last_value = val_str
                filled_series.loc[idx] = val_str
        
        df_filled[col] = filled_series
    
    return df_filled


def clean_single_sheet(df_raw: pd.DataFrame) -> Tuple[pd.DataFrame, dict]:
    """
    Structural clean for one tab:
    - Drop fully empty rows and columns
    - Detect header depth (1-3 rows)
    - Flatten and normalise headers
    - Remove header rows from the data
    - Flag exact duplicates (no removal)
    
    Returns:
        tuple: (cleaned_dataframe, metadata_dict)
    """
    df = df_raw.dropna(how="all").dropna(axis=1, how="all")
    
    rows_in = len(df)  # Track initial row count (after dropping empty rows/cols)
    
    metadata = {
        "duplicate_column_names_fixed": 0,
        "repeated_header_rows_dropped": 0,
        "totals_rows_dropped": 0,
        "totals_rows_flagged": 0,
        "header_row_index": 0,
        "header_depth_used": 1,
        "context_columns_filled": [],
        "rows_in": rows_in,
        "rows_out": 0,
        "info": [],  # Informational messages (successful fixes)
        "warnings": [],  # Things user should review
        "errors": [],  # Things that failed
    }

    if df.empty:
        metadata["rows_out"] = 0
        return df, metadata

    # Detect header depth (1-3 rows)
    header_idx, header_depth = detect_header_depth(df)
    metadata["header_row_index"] = header_idx
    metadata["header_depth_used"] = header_depth
    
    # Add warning if header depth is ambiguous (depth > 1 but might be wrong)
    if header_depth > 1:
        metadata["warnings"].append(f"Multi-row header detected (depth={header_depth}). Verify header structure.")
    
    # Flatten multi-row headers into single row
    flattened_header_row = flatten_multirow_headers(df, header_idx, header_depth)
    
    # Get data rows (skip all header rows)
    df_data = df.loc[header_idx + header_depth :].reset_index(drop=True)

    # Normalize the flattened headers
    normalized_headers, dup_col_count = normalise_headers(flattened_header_row)
    df_data.columns = normalized_headers
    metadata["duplicate_column_names_fixed"] = dup_col_count
    
    # Move successful fixes to INFO (not warnings)
    if dup_col_count > 0:
        metadata["info"].append(f"Fixed {dup_col_count} duplicate column name(s).")
    
    # Detect and remove repeated header rows
    df_data, repeated_header_count = detect_repeated_headers(df_data, normalized_headers)
    metadata["repeated_header_rows_dropped"] = repeated_header_count
    
    # Move successful fixes to INFO (not warnings)
    if repeated_header_count > 0:
        metadata["info"].append(f"Removed {repeated_header_count} repeated header row(s).")
    
    df_data = df_data.dropna(how="all")
    
    # Detect and flag/remove total rows
    if not df_data.empty:
        df_data, totals_flagged, totals_dropped = detect_total_rows(df_data)
        metadata["totals_rows_flagged"] = totals_flagged
        metadata["totals_rows_dropped"] = totals_dropped
        
        # Only warn if totals were flagged but not removed (needs review)
        if totals_flagged > 0:
            if totals_dropped > 0:
                metadata["info"].append(f"Removed {totals_dropped} total row(s).")
            if totals_flagged > totals_dropped:
                metadata["warnings"].append(f"{totals_flagged - totals_dropped} potential total row(s) flagged for review.")
    
    # Detect and fill down context columns
    if not df_data.empty:
        context_columns = detect_context_columns(df_data, normalized_headers)
        if context_columns:
            df_data = fill_down_context(df_data, context_columns)
            metadata["context_columns_filled"] = context_columns
            # Move successful fixes to INFO (not warnings)
            metadata["info"].append(f"Applied fill-down to {len(context_columns)} context column(s): {', '.join(context_columns[:3])}{'...' if len(context_columns) > 3 else ''}")

    if not df_data.empty:
        df_data["__possible_duplicate"] = df_data.duplicated(keep=False)
    else:
        df_data["__possible_duplicate"] = []
    
    # Track final row count
    metadata["rows_out"] = len(df_data)
    
    # Add warning if significant row reduction
    if rows_in > 0:
        reduction_ratio = (rows_in - metadata["rows_out"]) / rows_in
        if reduction_ratio > 0.5:
            metadata["warnings"].append(f"Significant row reduction: {rows_in} → {metadata['rows_out']} rows ({reduction_ratio*100:.1f}% reduction).")

    return df_data, metadata


# -----------------------------
# HELPER FUNCTIONS FOR SAFE VALUE HANDLING
# -----------------------------

def safe_isna(val: Any) -> bool:
    """
    Safely check if a value is NA, handling Series objects.
    Returns True if val is a Series (to skip it) or if it's NA.
    """
    if isinstance(val, pd.Series):
        return True  # Treat Series as NA to skip
    try:
        # pd.isna returns scalar bool for scalar values, Series for Series
        result = pd.isna(val)
        # If result is a Series, it means val was somehow a Series (shouldn't happen after isinstance check)
        if isinstance(result, pd.Series):
            return True  # Treat as NA
        # Otherwise, result is a scalar bool
        return bool(result)
    except (ValueError, TypeError):
        # If pd.isna fails, assume not NA
        return False

def safe_notna(val: Any) -> bool:
    """
    Safely check if a value is not NA, handling Series objects.
    Returns False if val is a Series (to skip it) or if it's NA.
    """
    if isinstance(val, pd.Series):
        return False  # Treat Series as NA to skip
    try:
        # pd.notna returns scalar bool for scalar values, Series for Series
        result = pd.notna(val)
        # If result is a Series, it means val was somehow a Series (shouldn't happen after isinstance check)
        if isinstance(result, pd.Series):
            return False  # Treat as NA
        # Otherwise, result is a scalar bool
        return bool(result)
    except (ValueError, TypeError):
        # If pd.notna fails, assume not NA
        return True

def safe_str_strip(val: Any) -> str:
    """
    Safely convert value to string and strip, handling Series objects.
    Returns empty string if val is a Series.
    """
    if isinstance(val, pd.Series):
        return ''
    try:
        return str(val).strip()
    except (ValueError, TypeError, AttributeError):
        return ''

# -----------------------------
# DATA SANITIZATION FOR SQL
# -----------------------------

def sanitize_for_sql(df: pd.DataFrame, type_analysis: Dict[str, Dict]) -> pd.DataFrame:
    """
    Sanitize DataFrame for SQL export:
    - Convert NULL values (empty strings, 'nan', 'None', 'null' text) to pd.NA
    - Standardize date formats to ISO 8601
    - Sanitize special characters (control characters, problematic Unicode)
    - Convert numeric columns to proper types where appropriate
    
    Args:
        df: DataFrame to sanitize
        type_analysis: Type analysis dict from analyze_column_types()
    
    Returns:
        Sanitized DataFrame ready for SQL export
    """
    if df.empty:
        return df
    
    df_sanitized = df.copy()
    
    # Remove internal columns from processing
    internal_cols = ['__possible_duplicate', '__is_total_row']
    data_cols = [c for c in df_sanitized.columns if c not in internal_cols]
    
    for col in data_cols:
        if col not in df_sanitized.columns:
            continue
        
        col_analysis = type_analysis.get(col, {})
        recommended_type = col_analysis.get('recommended_type', 'VARCHAR')
        
        # Convert NULL-like values to pd.NA
        # Replace empty strings, 'nan', 'None', 'null' (as text) with pd.NA
        df_sanitized[col] = df_sanitized[col].replace([
            '', 'nan', 'None', 'null', 'NULL', 'NaN', 'N/A', 'n/a'
        ], pd.NA)
        
        # Handle dates: convert to ISO 8601 format
        if recommended_type in ('DATE', 'TIMESTAMP_NTZ'):
            df_sanitized[col] = standardize_dates(df_sanitized[col], recommended_type)
        
        # Sanitize special characters (for all string columns)
        if recommended_type == 'VARCHAR':
            df_sanitized[col] = df_sanitized[col].apply(sanitize_string_value)
        
        # Convert numeric types (but keep as string for CSV - type conversion happens in SQL)
        # We'll validate ranges but keep as string to preserve precision
        if recommended_type in ('INTEGER', 'FLOAT'):
            df_sanitized[col] = validate_numeric_values(df_sanitized[col], recommended_type)
    
    return df_sanitized


def sanitize_string_value(val: Any) -> str:
    """
    Sanitize a single string value for SQL export.
    - Removes or replaces control characters
    - Normalizes problematic Unicode
    - Preserves NULL as pd.NA
    
    Args:
        val: Value to sanitize
    
    Returns:
        Sanitized string or pd.NA
    """
    # Use safe helper function to avoid Series boolean ambiguity
    if safe_isna(val):
        return pd.NA
    
    val_str = str(val)
    
    # Remove control characters except tab, newline, carriage return
    # Replace with space or remove based on character
    # Keep: \t (tab), \n (newline), \r (carriage return)
    # Remove: other control chars (0x00-0x1F except 0x09, 0x0A, 0x0D)
    sanitized = []
    for char in val_str:
        code = ord(char)
        if code < 32:  # Control character range
            if code in (9, 10, 13):  # Tab, newline, carriage return - keep
                sanitized.append(char)
            # else: skip other control characters
        elif code == 127:  # DEL character
            continue
        else:
            sanitized.append(char)
    
    result = ''.join(sanitized)
    
    # Return pd.NA if result is empty after sanitization
    if not result.strip():
        return pd.NA
    
    return result


def standardize_dates(series: pd.Series, date_type: str) -> pd.Series:
    """
    Standardize date values to ISO 8601 format.
    
    Args:
        series: Series with date values (as strings)
        date_type: 'DATE' or 'TIMESTAMP_NTZ'
    
    Returns:
        Series with standardized date strings or pd.NA
    """
    result = series.copy()
    date_only_formats = [
        '%Y-%m-%d',
        '%m/%d/%Y',
        '%d/%m/%Y',
        '%Y/%m/%d',
        '%d-%m-%Y',
        '%m-%d-%Y',
    ]
    timestamp_formats = [
        '%Y-%m-%d %H:%M:%S',
        '%m/%d/%Y %H:%M:%S',
        '%d/%m/%Y %H:%M:%S',
        '%Y-%m-%dT%H:%M:%S',
        '%Y/%m/%d %H:%M:%S',
        '%d/%m/%Y %H:%M:%S',
        '%Y-%m-%d %H:%M:%S.%f',
        '%Y-%m-%dT%H:%M:%S.%f',
    ]
    
    for idx in result.index:
        val = result.loc[idx]
        # Use safe helper functions to avoid Series boolean ambiguity
        if safe_isna(val):
            continue
        
        val_str = str(val).strip()
        if val_str == '' or val_str.lower() in ['nan', 'none', 'null']:
            result.loc[idx] = pd.NA
            continue
        
        # Try to parse and reformat
        parsed = None
        for fmt in (timestamp_formats if date_type == 'TIMESTAMP_NTZ' else date_only_formats):
            try:
                parsed = datetime.strptime(val_str, fmt)
                break
            except ValueError:
                continue
        
        # If parsing failed, try pandas' flexible parser
        if parsed is None:
            try:
                parsed = pd.to_datetime(val_str, errors='raise', infer_datetime_format=True)
            except (ValueError, TypeError):
                result.loc[idx] = pd.NA
                continue
        
        # Format to ISO 8601
        if date_type == 'TIMESTAMP_NTZ':
            result.loc[idx] = parsed.strftime('%Y-%m-%dT%H:%M:%S')
        else:  # DATE
            result.loc[idx] = parsed.strftime('%Y-%m-%d')
    
    return result


def validate_data_for_sql(df: pd.DataFrame, type_analysis: Dict[str, Dict]) -> Dict[str, List[str]]:
    """
    Validate data against SQL constraints and return warnings.
    
    Args:
        df: DataFrame to validate
        type_analysis: Type analysis dict from analyze_column_types()
    
    Returns:
        Dict with 'warnings' and 'errors' lists
    """
    validation_results = {
        'warnings': [],
        'errors': []
    }
    
    if df.empty:
        return validation_results
    
    internal_cols = ['__possible_duplicate', '__is_total_row']
    data_cols = [c for c in df.columns if c not in internal_cols]
    
    for col in data_cols:
        if col not in df.columns:
            continue
        
        col_analysis = type_analysis.get(col, {})
        recommended_type = col_analysis.get('recommended_type', 'VARCHAR')
        
        # Ensure we get a Series (handle duplicate column names or edge cases)
        try:
            col_data = df[col]
            if isinstance(col_data, pd.DataFrame):
                series = col_data.iloc[:, 0].dropna()
            else:
                series = pd.Series(col_data).dropna()
        except (KeyError, IndexError, TypeError):
            continue
        
        if not isinstance(series, pd.Series) or series.empty:
            continue
        
        # Validate VARCHAR length (only warn for very large values)
        if recommended_type == 'VARCHAR':
            max_len = 0
            long_values = []
            for val in series:
                val_str = str(val)
                byte_len = len(val_str.encode('utf-8'))
                if byte_len > max_len:
                    max_len = byte_len
                if byte_len > 16777216:  # Snowflake VARCHAR max
                    long_values.append((val_str[:50] + '...' if len(val_str) > 50 else val_str, byte_len))
            
            if long_values:
                validation_results['errors'].append(
                    f"Column '{col}': {len(long_values)} value(s) exceed Snowflake VARCHAR max (16MB)"
                )
            # Only warn for very large values (>100KB) - reduce noise
            elif max_len > 100000:
                validation_results['warnings'].append(
                    f"Column '{col}': Contains very large values (max {max_len:,} bytes) - consider VARIANT type for >16MB"
                )
        
        # Validate INTEGER range
        elif recommended_type == 'INTEGER':
            invalid_count = 0
            for val in series:
                try:
                    val_str = str(val).replace(',', '').replace('$', '').strip()
                    int_val = int(float(val_str))
                    # Check Python int range (Snowflake INTEGER is similar)
                    if int_val < -99999999999999999999999999999999999999 or int_val > 99999999999999999999999999999999999999:
                        invalid_count += 1
                except (ValueError, OverflowError):
                    invalid_count += 1
            
            if invalid_count > 0:
                validation_results['warnings'].append(
                    f"Column '{col}': {invalid_count} value(s) cannot be converted to INTEGER"
                )
        
        # Validate FLOAT range
        elif recommended_type == 'FLOAT':
            invalid_count = 0
            for val in series:
                try:
                    val_str = str(val).replace(',', '').replace('$', '').strip()
                    float_val = float(val_str)
                    if not math.isfinite(float_val):
                        invalid_count += 1
                except (ValueError, OverflowError):
                    invalid_count += 1
            
            if invalid_count > 0:
                validation_results['warnings'].append(
                    f"Column '{col}': {invalid_count} value(s) cannot be converted to FLOAT"
                )
        
        # Validate DATE/TIMESTAMP format (only warn if significant portion is invalid)
        elif recommended_type in ('DATE', 'TIMESTAMP_NTZ'):
            invalid_count = 0
            total_count = len(series)
            for val in series:
                val_str = str(val).strip()
                try:
                    # Try to parse as date
                    pd.to_datetime(val_str, errors='raise')
                except (ValueError, TypeError):
                    invalid_count += 1
            
            # Only warn if >10% of values are invalid (reduce noise for edge cases)
            if invalid_count > 0 and (invalid_count / total_count) > 0.1:
                validation_results['warnings'].append(
                    f"Column '{col}': {invalid_count} of {total_count} value(s) ({invalid_count/total_count*100:.1f}%) may not be valid {recommended_type} format"
                )
    
    return validation_results


def validate_numeric_values(series: pd.Series, numeric_type: str) -> pd.Series:
    """
    Validate numeric values and keep as string (for CSV export).
    Invalid values become pd.NA.
    
    Args:
        series: Series with numeric values (as strings)
        numeric_type: 'INTEGER' or 'FLOAT'
    
    Returns:
        Series with validated numeric strings or pd.NA
    """
    result = series.copy()
    
    for idx in result.index:
        val = result.loc[idx]
        # Use safe helper functions to avoid Series boolean ambiguity
        if safe_isna(val):
            continue
        
        val_str = str(val).strip()
        if val_str == '' or val_str.lower() in ['nan', 'none', 'null']:
            result.loc[idx] = pd.NA
            continue
        
        # Try to parse as number
        try:
            # Remove common formatting
            clean_val = val_str.replace(',', '').replace('$', '').replace('%', '').strip()
            
            if numeric_type == 'INTEGER':
                int_val = int(float(clean_val))  # Convert via float to handle "123.0"
                # Check range (Snowflake INTEGER is -99999999999999999999999999999999999999 to 99999999999999999999999999999999999999)
                # For practical purposes, we'll use Python int limits
                result.loc[idx] = str(int_val)
            else:  # FLOAT
                float_val = float(clean_val)
                # Check for infinity/NaN
                if math.isfinite(float_val):
                    result.loc[idx] = str(float_val)
                else:
                    result.loc[idx] = pd.NA
        except (ValueError, OverflowError):
            result.loc[idx] = pd.NA
    
    return result


# -----------------------------
# TYPE ANALYSIS
# -----------------------------

def analyze_column_types(df: pd.DataFrame) -> Dict[str, Dict]:
    """
    Analyze each column for type parseability and recommend Snowflake types.
    Keeps data as string - only provides recommendations.
    
    Returns:
        dict: {column_name: {
            'pct_int': float,
            'pct_float': float,
            'pct_date': float,
            'recommended_type': str,
            'sample_values': list
        }}
    """
    type_analysis = {}
    
    if df.empty:
        return type_analysis
    
    cols = [c for c in df.columns if c != "__possible_duplicate"]
    
    for col in cols:
        # Ensure we get a Series (handle duplicate column names or edge cases)
        try:
            col_data = df[col]
            if isinstance(col_data, pd.DataFrame):
                series = col_data.iloc[:, 0].dropna().astype(str)
            else:
                series = pd.Series(col_data).dropna().astype(str)
        except (KeyError, IndexError, TypeError):
            continue
        
        if not isinstance(series, pd.Series) or series.empty:
            type_analysis[col] = {
                'pct_int': 0.0,
                'pct_float': 0.0,
                'pct_date': 0.0,
                'recommended_type': 'VARCHAR',
                'sample_values': []
            }
            continue
        
        total = len(series)
        int_count = 0
        float_count = 0
        date_count = 0
        has_timestamp = False  # Track if any dates have time components
        
        # Sample first 5 non-empty values for display
        sample_values = series.head(5).tolist()
        
        # Date formats - separate date-only and timestamp formats
        date_only_formats = [
            '%Y-%m-%d',
            '%m/%d/%Y',
            '%d/%m/%Y',
        ]
        timestamp_formats = [
            '%Y-%m-%d %H:%M:%S',
            '%m/%d/%Y %H:%M:%S',
            '%d/%m/%Y %H:%M:%S',
            '%Y-%m-%dT%H:%M:%S',
        ]
        
        for val in series:
            val_str = str(val).strip()
            if val_str == "" or val_str.lower() in ["nan", "none", "null"]:
                continue
            
            # Try integer parsing
            try:
                int(val_str)
                int_count += 1
                continue
            except (ValueError, OverflowError):
                pass
            
            # Try float parsing
            try:
                float(val_str)
                float_count += 1
                continue
            except (ValueError, OverflowError):
                pass
            
            # Try date parsing (date-only formats first)
            date_parsed = False
            for fmt in date_only_formats:
                try:
                    datetime.strptime(val_str, fmt)
                    date_count += 1
                    date_parsed = True
                    break
                except ValueError:
                    continue
            
            # Try timestamp formats if date-only didn't work
            if not date_parsed:
                for fmt in timestamp_formats:
                    try:
                        datetime.strptime(val_str, fmt)
                        date_count += 1
                        has_timestamp = True
                        date_parsed = True
                        break
                    except ValueError:
                        continue
        
        pct_int = (int_count / total * 100) if total > 0 else 0.0
        pct_float = (float_count / total * 100) if total > 0 else 0.0
        pct_date = (date_count / total * 100) if total > 0 else 0.0
        
        # Recommend Snowflake type
        if pct_date >= 80:
            recommended_type = "TIMESTAMP_NTZ" if has_timestamp else "DATE"
        elif pct_int >= 90:
            recommended_type = "INTEGER"
        elif pct_float >= 80:
            recommended_type = "FLOAT"
        else:
            recommended_type = "VARCHAR"
        
        type_analysis[col] = {
            'pct_int': round(pct_int, 1),
            'pct_float': round(pct_float, 1),
            'pct_date': round(pct_date, 1),
            'recommended_type': recommended_type,
            'sample_values': sample_values[:3]  # Keep only 3 for display
        }
    
    return type_analysis


# -----------------------------
# SQL STATEMENT GENERATION
# -----------------------------

def generate_create_table_statements(
    cleaned_sheets: Dict[str, pd.DataFrame],
    sheet_metadata: Dict[str, Dict],
    type_analysis_df: pd.DataFrame
) -> str:
    """
    Generate CREATE TABLE SQL statements for all cleaned sheets.
    Uses Snowflake SQL syntax.
    
    Args:
        cleaned_sheets: Dict of sheet_name -> DataFrame
        sheet_metadata: Dict of sheet_name -> metadata dict
        type_analysis_df: DataFrame with type analysis
    
    Returns:
        String containing all CREATE TABLE statements
    """
    sql_statements = []
    sql_statements.append("-- SQL DDL statements for Snowflake")
    sql_statements.append("-- Generated automatically from normalized spreadsheet")
    sql_statements.append("-- Review and adjust data types as needed")
    sql_statements.append("")
    
    # Create a lookup for type recommendations
    type_lookup = {}
    if not type_analysis_df.empty and 'Tab name' in type_analysis_df.columns:
        for _, row in type_analysis_df.iterrows():
            tab_name = row.get('Tab name', '')
            col_name = row.get('Column name', '')
            rec_type = row.get('Recommended Snowflake type', 'VARCHAR')
            if tab_name and col_name:
                if tab_name not in type_lookup:
                    type_lookup[tab_name] = {}
                type_lookup[tab_name][col_name] = rec_type
    
    for sheet_name, df in cleaned_sheets.items():
        if df.empty:
            continue
        
        # Sanitize table name (same rules as column names)
        table_name = sanitize_identifier(sheet_name)
        if not table_name or table_name[0].isdigit():
            table_name = f"table_{table_name}"
        if table_name in SQL_RESERVED_WORDS:
            table_name = f"{table_name}_tbl"
        if len(table_name) > 255:
            table_name = table_name[:252] + "..."
        
        sql_statements.append(f"-- Table: {sheet_name}")
        sql_statements.append(f"CREATE TABLE {table_name} (")
        
        # Get type recommendations for this sheet
        sheet_types = type_lookup.get(sheet_name, {})
        
        columns = []
        for col in df.columns:
            if col in ['__possible_duplicate', '__is_total_row']:
                continue
            
            # Get recommended type
            rec_type = sheet_types.get(col, 'VARCHAR')
            
            # Determine size for VARCHAR
            if rec_type == 'VARCHAR':
                # Calculate max length in this column
                max_len = 0
                for val in df[col].dropna():
                    val_str = str(val)
                    # Estimate UTF-8 byte length (rough approximation)
                    byte_len = len(val_str.encode('utf-8'))
                    max_len = max(max_len, byte_len)
                
                # Add 20% buffer and round up to reasonable sizes
                if max_len == 0:
                    varchar_size = 16777216  # Snowflake default VARCHAR max
                elif max_len < 100:
                    varchar_size = 255
                elif max_len < 500:
                    varchar_size = 500
                elif max_len < 2000:
                    varchar_size = 2000
                elif max_len < 10000:
                    varchar_size = 10000
                else:
                    varchar_size = 16777216  # Snowflake max
                
                type_def = f"VARCHAR({varchar_size})"
            else:
                type_def = rec_type
            
            # Sanitize column name (should already be done, but double-check)
            col_safe = sanitize_identifier(col)
            if not col_safe or col_safe[0].isdigit():
                col_safe = f"col_{col_safe}"
            if col_safe in SQL_RESERVED_WORDS:
                col_safe = f"{col_safe}_col"
            
            columns.append(f"    {col_safe} {type_def}")
        
        sql_statements.append(",\n".join(columns))
        sql_statements.append(");")
        sql_statements.append("")
    
    return "\n".join(sql_statements)


def sanitize_identifier(identifier: str) -> str:
    """
    Sanitize an identifier (table or column name) for SQL.
    Applies same rules as normalise_headers but for a single identifier.
    
    Args:
        identifier: Identifier to sanitize
    
    Returns:
        SQL-safe identifier
    """
    if not identifier:
        return "unnamed"
    
    # Normalize: lowercase, underscores, remove special chars
    clean = (
        str(identifier)
        .strip()
        .lower()
        .replace(' ', '_')
    )
    clean = re.sub(r'[^\w]', '', clean)
    
    if not clean or clean == "nan":
        return "unnamed"
    
    # Fix numeric start
    if clean[0].isdigit():
        clean = f"col_{clean}"
    
    # Fix reserved words
    if clean in SQL_RESERVED_WORDS:
        clean = f"{clean}_col"
    
    # Fix length
    if len(clean) > 255:
        clean = clean[:252] + "..."
    
    return clean


# -----------------------------
# CANDIDATE KEYS
# -----------------------------

def find_candidate_keys(df: pd.DataFrame) -> list:
    """
    Find columns that look like keys (SKU, ID, etc).
    Criteria:
    - Column not empty
    - Fill ratio >= 0.7
    - Uniqueness ratio >= 0.7
    """
    if df.empty:
        return []

    n_rows = len(df)
    candidates = []
    cols = [c for c in df.columns if c != "__possible_duplicate"]

    for col in cols:
        # Ensure we get a Series (handle duplicate column names or edge cases)
        try:
            col_data = df[col]
            if isinstance(col_data, pd.DataFrame):
                series = col_data.iloc[:, 0].dropna().astype(str)
            else:
                series = pd.Series(col_data).dropna().astype(str)
        except (KeyError, IndexError, TypeError):
            continue
        
        if not isinstance(series, pd.Series) or series.empty:
            continue

        fill_ratio = len(series) / n_rows
        uniqueness_ratio = series.nunique() / len(series)

        if fill_ratio >= 0.7 and uniqueness_ratio >= 0.7:
            candidates.append(col)

    return candidates


# -----------------------------
# JOB ID GENERATION
# -----------------------------

def generate_job_id() -> str:
    """
    Generate a unique job ID for this normalization run.
    Returns a short UUID string.
    """
    return str(uuid.uuid4())[:8]


# -----------------------------
# MAIN PIPELINE
# -----------------------------

def normalize_spreadsheet(
    input_path: Path,
    output_format: str,  # "1" = Excel, "2" = CSV, "3" = Both
    output_dir: Optional[Path] = None
) -> Dict[str, Any]:
    """
    Normalize spreadsheet and return results.
    
    Args:
        input_path: Path to input Excel file
        output_format: "1" = Excel only, "2" = CSV only, "3" = Both
        output_dir: Optional directory for output files. If None, uses current directory.
    
    Returns:
        Dictionary with:
        - 'job_id': str
        - 'cleaned_sheets': Dict[str, pd.DataFrame]
        - 'meta_df': pd.DataFrame
        - 'type_analysis_df': pd.DataFrame
        - 'sheet_metadata': Dict[str, Dict]
        - 'excel_output_path': Optional[Path]
        - 'csv_output_path': Optional[Path]
        - 'status': str ("success", "partial", "error")
        - 'errors': List[str]
        - 'warnings': List[str]
    """
    job_id = generate_job_id()
    
    # Set output directory
    if output_dir is None:
        output_dir = Path(".")
    else:
        output_dir.mkdir(parents=True, exist_ok=True)
    
    # Determine temp file format based on input
    if input_path.suffix.lower() == '.xls':
        # For .xls files, read directly (merged cell expansion not supported)
        temp_file = input_path
    else:
        # For .xlsx files, expand merged cells into temp file
        temp_file = output_dir / f"_temp_unmerged_{job_id}.xlsx"
        expand_merged_cells(input_path, temp_file)
    
    excel_output_file = output_dir / f"clean_{input_path.stem}{input_path.suffix}"
    csv_output_file = output_dir / f"clean_{input_path.stem}_ALL.csv"

    # Read and clean each sheet (with region detection)
    xl = pd.ExcelFile(temp_file)
    cleaned_sheets = {}
    sheet_metadata = {}  # Store metadata for each virtual sheet
    meta_rows = []
    all_type_analysis = []  # Collect detailed type analysis for TYPE_ANALYSIS sheet
    all_errors = []
    all_warnings = []
    all_info = []  # Collect informational messages (successful fixes)

    for sheet_name in xl.sheet_names:
        try:
            df_raw = pd.read_excel(
                xl, sheet_name=sheet_name, header=None, dtype=str
            )
            
            # Detect table regions within this sheet
            regions = detect_table_regions(df_raw)
            
            # Process each region as a separate "virtual sheet"
            for region_idx, region in enumerate(regions, 1):
                # Extract region from raw dataframe
                region_df_raw = df_raw.loc[region['min_row']:region['max_row'], 
                                          region['min_col']:region['max_col']].copy()
                
                # Create virtual sheet name
                if len(regions) > 1:
                    virtual_sheet_name = f"{sheet_name}__table{region_idx:02d}"
                else:
                    virtual_sheet_name = sheet_name

                # Process this region
                try:
                    df_clean, clean_metadata = clean_single_sheet(region_df_raw)
                except ValueError as e:
                    # Catch "ambiguous truth value" errors and provide better error message
                    if "truth value" in str(e).lower() or "ambiguous" in str(e).lower():
                        error_msg = f"Data structure issue in sheet '{sheet_name}': {str(e)}. This may be due to duplicate columns or unusual data structure. The sheet will be skipped."
                        all_errors.append(f"{virtual_sheet_name}: {error_msg}")
                        clean_metadata = {
                            "duplicate_column_names_fixed": 0,
                            "repeated_header_rows_dropped": 0,
                            "totals_rows_dropped": 0,
                            "totals_rows_flagged": 0,
                            "header_row_index": 0,
                            "header_depth_used": 1,
                            "context_columns_filled": [],
                            "rows_in": len(region_df_raw),
                            "rows_out": 0,
                            "info": [],
                            "warnings": [],
                            "errors": [error_msg],
                        }
                        df_clean = pd.DataFrame()  # Empty DataFrame for failed sheets
                    else:
                        raise  # Re-raise if it's a different ValueError
                
                # Add table bounds to metadata
                clean_metadata['table_bounds'] = region['bounds']
                
                candidate_keys = find_candidate_keys(df_clean)
                dup_count = int(df_clean["__possible_duplicate"].sum()) if "__possible_duplicate" in df_clean.columns else 0
                dup_col_count = clean_metadata.get("duplicate_column_names_fixed", 0)
                repeated_header_count = clean_metadata.get("repeated_header_rows_dropped", 0)
                totals_flagged = clean_metadata.get("totals_rows_flagged", 0)
                totals_dropped = clean_metadata.get("totals_rows_dropped", 0)
                header_row_index = clean_metadata.get("header_row_index", 0)
                header_depth_used = clean_metadata.get("header_depth_used", 1)
                context_columns_filled = clean_metadata.get("context_columns_filled", [])
                table_bounds = clean_metadata.get("table_bounds", "")
                rows_in = clean_metadata.get("rows_in", 0)
                rows_out = clean_metadata.get("rows_out", 0)
                warnings = clean_metadata.get("warnings", [])
                errors = clean_metadata.get("errors", [])
                info = clean_metadata.get("info", [])
                
                # Collect info, warnings, and errors separately
                all_info.extend([f"{virtual_sheet_name}: {i}" for i in info])
                all_warnings.extend([f"{virtual_sheet_name}: {w}" for w in warnings])
                all_errors.extend([f"{virtual_sheet_name}: {e}" for e in errors])
                
                # Add warning if multiple tables detected
                if len(regions) > 1:
                    warnings.append(f"Sheet split into {len(regions)} table region(s).")
                
                # Analyze column types
                type_analysis = analyze_column_types(df_clean)
                
                # Validate data for SQL constraints
                validation_results = validate_data_for_sql(df_clean, type_analysis)
                if validation_results['warnings']:
                    clean_metadata['warnings'].extend(validation_results['warnings'])
                    all_warnings.extend([f"{virtual_sheet_name}: {w}" for w in validation_results['warnings']])
                if validation_results['errors']:
                    clean_metadata['errors'].extend(validation_results['errors'])
                    all_errors.extend([f"{virtual_sheet_name}: {e}" for e in validation_results['errors']])
                
                # Create type summary string for META sheet
                type_summary_parts = []
                type_counts = {}
                for col, analysis in type_analysis.items():
                    rec_type = analysis['recommended_type']
                    type_counts[rec_type] = type_counts.get(rec_type, 0) + 1
                
                type_summary = ", ".join([f"{k}({v})" for k, v in sorted(type_counts.items())])
                if not type_summary:
                    type_summary = "N/A"

                # Sanitize data for SQL export (dates, NULLs, special chars)
                # This ensures both Excel and CSV outputs are SQL-ready
                df_clean = sanitize_for_sql(df_clean, type_analysis)

                cleaned_sheets[virtual_sheet_name] = df_clean
                # Store full type analysis in metadata for potential future use
                clean_metadata['type_analysis'] = type_analysis
                clean_metadata['source_tab'] = sheet_name
                clean_metadata['source_table_id'] = f"table{region_idx:02d}" if len(regions) > 1 else ""
                sheet_metadata[virtual_sheet_name] = clean_metadata
                
                # Collect detailed type analysis for TYPE_ANALYSIS sheet
                for col_name, analysis in type_analysis.items():
                    all_type_analysis.append({
                        "Tab name": virtual_sheet_name,
                        "Column name": col_name,
                        "Recommended Snowflake type": analysis['recommended_type'],
                        "% parseable as INT": analysis['pct_int'],
                        "% parseable as FLOAT": analysis['pct_float'],
                        "% parseable as DATE": analysis['pct_date'],
                        "Sample values": ", ".join([str(v) for v in analysis['sample_values']])
                    })

                meta_rows.append(
                    {
                        "Tab name": virtual_sheet_name,
                        "Source tab": sheet_name,
                        "Source table ID": clean_metadata['source_table_id'],
                        "Clean status": "OK",
                        "Rows in": rows_in,
                        "Rows out": rows_out,
                        "Rows after clean": df_clean.shape[0],
                        "Columns after clean": df_clean.shape[1],
                        "Header row index": header_row_index,
                        "Header depth used": header_depth_used,
                        "Table bounds": table_bounds,
                        "Possible key columns": ", ".join(candidate_keys),
                        "Exact duplicate rows": dup_count,
                        "Duplicate column names fixed": dup_col_count,
                        "Repeated header rows dropped": repeated_header_count,
                        "Totals rows flagged": totals_flagged,
                        "Totals rows dropped": totals_dropped,
                        "Context columns filled": ", ".join(context_columns_filled) if context_columns_filled else "",
                        "Type recommendations": type_summary,
                        "Warnings": "; ".join(warnings) if warnings else "",
                        "Errors": "; ".join(errors) if errors else "",
                        "Comments / error": "",
                    }
                )

        except Exception as e:
            error_msg = str(e)
            all_errors.append(f"{sheet_name}: {error_msg}")
            meta_rows.append(
                {
                    "Tab name": sheet_name,
                    "Source tab": sheet_name,
                    "Source table ID": "",
                    "Clean status": "ERROR",
                    "Rows in": 0,
                    "Rows out": 0,
                    "Rows after clean": 0,
                    "Columns after clean": 0,
                    "Header row index": 0,
                    "Header depth used": 0,
                    "Table bounds": "",
                    "Possible key columns": "",
                    "Exact duplicate rows": 0,
                    "Duplicate column names fixed": 0,
                    "Repeated header rows dropped": 0,
                    "Totals rows flagged": 0,
                    "Totals rows dropped": 0,
                    "Context columns filled": "",
                    "Type recommendations": "",
                    "Warnings": "",
                    "Errors": error_msg,
                    "Comments / error": error_msg,
                }
            )

    meta_df = pd.DataFrame(meta_rows)

    # Data quality flag
    def flag(row):
        # ERROR if exception occurred
        if row["Clean status"] != "OK":
            return "ERROR - Check tab"
        
        # Check for errors
        errors = str(row.get("Errors", "") if "Errors" in row.index else "")
        if errors and errors.strip() and errors.lower() != "nan":
            return "ERROR - Check tab"
        
        # REVIEW if warnings exist or significant issues
        warnings = str(row.get("Warnings", "") if "Warnings" in row.index else "")
        has_warnings = warnings and warnings.strip() and warnings.lower() != "nan"
        
        if has_warnings:
            return "REVIEW - Warnings present"
        
        if row.get("Exact duplicate rows", 0) >= 100:
            return "REVIEW - High duplicate count"
        
        if row.get("Totals rows dropped", 0) > 0:
            return "REVIEW - Totals removed"
        
        if row.get("Repeated header rows dropped", 0) > 0:
            return "REVIEW - Repeated headers removed"
        
        # OK only if clean + no warnings
        return "OK"

    meta_df["Data quality flag"] = meta_df.apply(flag, axis=1)

    # Excel output
    excel_output_path = None
    if output_format in ("1", "3"):
        with pd.ExcelWriter(excel_output_file, engine="openpyxl") as writer:
            for name, df in cleaned_sheets.items():
                safe_name = name[:31]
                df.to_excel(writer, sheet_name=safe_name, index=False)

            meta_df.to_excel(writer, sheet_name="META", index=False)
            
            # Add detailed type analysis sheet
            if all_type_analysis:
                type_analysis_df = pd.DataFrame(all_type_analysis)
                type_analysis_df.to_excel(writer, sheet_name="TYPE_ANALYSIS", index=False)

        excel_output_path = excel_output_file

    # CSV output (single file with all tabs)
    csv_output_path = None
    if output_format in ("2", "3"):
        all_frames = []

        for virtual_sheet_name, df in cleaned_sheets.items():
            if df.empty:
                continue

            df_copy = df.copy()

            # Remove duplicate column names inside this tab
            df_copy = df_copy.loc[:, ~df_copy.columns.duplicated()]

            # Get metadata for this sheet
            metadata = sheet_metadata.get(virtual_sheet_name, {})
            source_tab = metadata.get('source_tab', virtual_sheet_name)
            source_table_id = metadata.get('source_table_id', '')
            
            # Data is already sanitized before storing in cleaned_sheets
            # No need to sanitize again here

            # Add source_tab and source_table_id columns
            df_copy.insert(0, "source_tab", source_tab)
            if source_table_id:
                df_copy.insert(1, "source_table_id", source_table_id)

            # Reset index to avoid any index confusion
            df_copy = df_copy.reset_index(drop=True)

            all_frames.append(df_copy)

        if all_frames:
            combined = pd.concat(all_frames, ignore_index=True, sort=False)
            
            # Export CSV with SQL-optimized settings
            # - UTF-8 encoding (required for Snowflake)
            # - Proper quoting for fields with commas/quotes/newlines
            # - Explicit NULL handling (empty strings become empty, pd.NA becomes empty)
            combined.to_csv(
                csv_output_file,
                index=False,
                encoding='utf-8',
                quoting=csv.QUOTE_MINIMAL,  # Quote only when necessary
                escapechar=None,  # Use double quotes for escaping
                doublequote=True,  # Double quotes within quoted fields
                na_rep='',  # Represent NULL/NA as empty string (SQL standard)
                lineterminator='\n'  # Unix line endings (Snowflake compatible) - FIXED: was line_terminator
            )
            csv_output_path = csv_output_file

    # Determine status
    if all_errors:
        status = "error"
    elif all_warnings:
        status = "partial"
    else:
        status = "success"

    # Clean up temp file if we created it
    if temp_file != input_path and temp_file.exists():
        try:
            temp_file.unlink()
        except:
            pass

    # Create type_analysis_df
    type_analysis_df = pd.DataFrame(all_type_analysis) if all_type_analysis else pd.DataFrame()

    # Generate SQL CREATE TABLE statements
    sql_output_path = None
    if cleaned_sheets:
        sql_output_file = output_dir / f"clean_{input_path.stem}_CREATE_TABLES.sql"
        try:
            sql_statements = generate_create_table_statements(
                cleaned_sheets,
                sheet_metadata,
                type_analysis_df
            )
            with open(sql_output_file, 'w', encoding='utf-8') as f:
                f.write(sql_statements)
            sql_output_path = sql_output_file
        except Exception as e:
            all_warnings.append(f"Could not generate SQL file: {str(e)}")

    return {
        'job_id': job_id,
        'cleaned_sheets': cleaned_sheets,
        'meta_df': meta_df,
        'type_analysis_df': type_analysis_df,
        'sheet_metadata': sheet_metadata,
        'excel_output_path': excel_output_path,
        'csv_output_path': csv_output_path,
        'sql_output_path': sql_output_path,
        'status': status,
        'info': all_info,  # Informational messages (successful fixes)
        'warnings': all_warnings,  # Things user should review
        'errors': all_errors,  # Things that failed
    }


def main():
    """
    CLI entry point - preserved for backward compatibility.
    Uses interactive file selection and format choice.
    """
    input_file = choose_input_file()
    output_choice = choose_output_format()
    
    print(f"\nProcessing: {input_file.name}")
    
    # Use the refactored function
    results = normalize_spreadsheet(input_file, output_choice)
    
    # Print results summary
    if results['excel_output_path']:
        print(f"\nExcel saved → {results['excel_output_path']}")
    if results['csv_output_path']:
        print(f"CSV saved → {results['csv_output_path']}")
    
    if results['errors']:
        print(f"\nErrors encountered: {len(results['errors'])}")
    if results['warnings']:
        print(f"Warnings: {len(results['warnings'])}")
    
    print("\nDone.")


if __name__ == "__main__":
    main()
